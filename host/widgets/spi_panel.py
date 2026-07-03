"""SPI 寄存器写入面板, 支持命名多预设."""

import json, os
from datetime import datetime
from typing import List, Tuple

import openpyxl
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QSpinBox, QComboBox, QTableWidget, QTableWidgetItem,
    QGroupBox, QMessageBox, QCheckBox, QInputDialog, QFileDialog,
)
from PyQt6.QtCore import Qt, pyqtSignal, QTimer

from comm.protocol import build_spi_config, build_spi_reg_writes
from comm.serial_link import SerialLink

_PRESET_PATH = os.path.join(os.path.expanduser("~"), ".h723_spi_presets.json")

SEND_COL_CHECK  = 0
SEND_COL_ADDR   = 1
SEND_COL_DATA   = 2
PRESET_COL_ADDR = 4
PRESET_COL_DATA = 5
NUM_ROWS        = 8


def _default_preset():
    return {str(r): ["00", "00"] for r in range(NUM_ROWS)}


class SpiPanel(QWidget):
    log_signal = pyqtSignal(str)

    def __init__(self, link: SerialLink, parent=None):
        super().__init__(parent)
        self.link = link
        self._presets_db = {}
        self._setup_ui()
        self._load_presets()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        # --- SPI 配置 ---
        cfg_box = QGroupBox("SPI 配置")
        cfg_layout = QHBoxLayout(cfg_box)
        cfg_layout.addWidget(QLabel("CPOL:"))
        self.cb_cpol = QComboBox()
        self.cb_cpol.addItems(["0", "1"])
        cfg_layout.addWidget(self.cb_cpol)

        cfg_layout.addWidget(QLabel("CPHA:"))
        self.cb_cpha = QComboBox()
        self.cb_cpha.addItems(["0", "1"])
        cfg_layout.addWidget(self.cb_cpha)

        cfg_layout.addWidget(QLabel("分频器:"))
        self.cb_prescaler = QComboBox()
        for i, label in enumerate(["/2","/4","/8","/16","/32","/64","/128","/256"]):
            self.cb_prescaler.addItem(label, i)
        self.cb_prescaler.setCurrentIndex(5)
        cfg_layout.addWidget(self.cb_prescaler)

        cfg_layout.addWidget(QLabel("位宽:"))
        self.cb_dwidth = QComboBox()
        self.cb_dwidth.addItem("16-bit", 16)
        cfg_layout.addWidget(self.cb_dwidth)

        self.btn_cfg = QPushButton("应用配置")
        self.btn_cfg.clicked.connect(self._send_config)
        cfg_layout.addWidget(self.btn_cfg)
        layout.addWidget(cfg_box)

        # --- 引脚映射 ---
        pin_box = QGroupBox("引脚映射")
        pin_layout = QHBoxLayout(pin_box)
        pin_layout.setSpacing(20)
        for pin, func in [("PB12","CS"),("PB13","CLK"),("PB14","MISO"),("PB15","MOSI")]:
            lbl = QLabel(f"<b>{pin}</b><br>{func}")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            pin_layout.addWidget(lbl)
        layout.addWidget(pin_box)

        # --- 寄存器表 ---
        reg_box = QGroupBox("寄存器写入  (左侧 = 发送值, 右侧 = 预设)")
        reg_layout = QVBoxLayout(reg_box)

        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["✓","地址","数据","","预设地址","预设数据"])
        self.table.setRowCount(NUM_ROWS)
        self.table.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        self.table.horizontalHeader().setSectionResizeMode(0, self.table.horizontalHeader().ResizeMode.Fixed)
        self.table.setColumnWidth(0, 30)

        self._row_checks = []
        for r in range(NUM_ROWS):
            chk = QCheckBox()
            chk.setChecked(True)
            self._row_checks.append(chk)
            self.table.setCellWidget(r, SEND_COL_CHECK, chk)
            for c in (SEND_COL_ADDR, SEND_COL_DATA, PRESET_COL_ADDR, PRESET_COL_DATA):
                item = QTableWidgetItem("00")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(r, c, item)
            sep = QTableWidgetItem("")
            sep.setFlags(Qt.ItemFlag.NoItemFlags)
            sep.setBackground(Qt.GlobalColor.lightGray)
            self.table.setItem(r, 3, sep)
        self.table.setColumnWidth(3, 30)
        self.table.horizontalHeader().setStretchLastSection(True)
        reg_layout.addWidget(self.table)

        # --- 预设选择器 ---
        ps_layout = QHBoxLayout()
        ps_layout.addWidget(QLabel("预设:"))
        self.cb_preset = QComboBox()
        self.cb_preset.setEditable(True)
        self.cb_preset.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.cb_preset.setPlaceholderText("输入名称创建新预设...")
        self.cb_preset.setMinimumWidth(140)
        self.cb_preset.activated.connect(self._on_preset_activated)
        ps_layout.addWidget(self.cb_preset)

        self.btn_save_preset = QPushButton("保存预设")
        self.btn_save_preset.clicked.connect(self._save_preset)
        ps_layout.addWidget(self.btn_save_preset)

        self.btn_rename_preset = QPushButton("重命名")
        self.btn_rename_preset.clicked.connect(self._rename_preset)
        ps_layout.addWidget(self.btn_rename_preset)

        self.btn_delete_preset = QPushButton("删除")
        self.btn_delete_preset.clicked.connect(self._delete_preset)
        ps_layout.addWidget(self.btn_delete_preset)

        self.btn_export = QPushButton("导出Excel")
        self.btn_export.clicked.connect(self._export_excel)
        ps_layout.addWidget(self.btn_export)

        self.btn_import = QPushButton("导入Excel")
        self.btn_import.clicked.connect(self._import_excel)
        ps_layout.addWidget(self.btn_import)
        ps_layout.addStretch()
        reg_layout.addLayout(ps_layout)

        # --- 操作按钮 ---
        btn_layout = QHBoxLayout()
        self.btn_fill = QPushButton("填至左侧")
        self.btn_fill.clicked.connect(self._fill_from_presets)
        btn_layout.addWidget(self.btn_fill)
        btn_layout.addStretch()
        reg_layout.addLayout(btn_layout)

        # --- 发送控制 ---
        send_layout = QHBoxLayout()
        self.btn_send = QPushButton("发送")
        self.btn_send.clicked.connect(self._send_regs)
        send_layout.addWidget(self.btn_send)

        self.chk_burst = QCheckBox("持续发送")
        self.chk_burst.stateChanged.connect(self._toggle_burst)
        send_layout.addWidget(self.chk_burst)

        self.chk_nonstop = QCheckBox("不间断")
        self.chk_nonstop.stateChanged.connect(self._toggle_nonstop)
        send_layout.addWidget(self.chk_nonstop)

        send_layout.addWidget(QLabel("间隔(ms):"))
        self.sb_interval = QSpinBox()
        self.sb_interval.setRange(1, 1000)
        self.sb_interval.setValue(5)
        send_layout.addWidget(self.sb_interval)
        reg_layout.addLayout(send_layout)

        layout.addWidget(reg_box)
        layout.addStretch()

        self._burst_timer = QTimer(self)
        self._burst_timer.timeout.connect(self._send_regs)

    # ----------------------------------------------------------------
    # 预设持久化
    # ----------------------------------------------------------------

    def _load_presets(self):
        if not os.path.exists(_PRESET_PATH):
            self._presets_db = {}
            return
        try:
            with open(_PRESET_PATH, "r", encoding="utf-8") as f:
                raw = json.load(f)
            if isinstance(raw, dict):
                all_rows = all(isinstance(v, list) for v in raw.values())
                if raw and all_rows:
                    self._presets_db = {"Default": raw}
                    self._save_preset_db()
                else:
                    self._presets_db = raw
            else:
                self._presets_db = {"Default": raw}
                self._save_preset_db()
        except Exception:
            self._presets_db = {}

        self.cb_preset.blockSignals(True)
        self.cb_preset.clear()
        names = sorted(self._presets_db.keys())
        self.cb_preset.addItems(names)
        if names:
            self.cb_preset.setCurrentIndex(0)
        self.cb_preset.blockSignals(False)

        if names:
            self._show_preset(names[0])

    def _save_preset_db(self):
        try:
            with open(_PRESET_PATH, "w", encoding="utf-8") as f:
                json.dump(self._presets_db, f, ensure_ascii=False, indent=2)
        except Exception as e:
            QMessageBox.warning(self, "保存失败", str(e))

    def _gather_right_side(self):
        data = {}
        for r in range(NUM_ROWS):
            ai = self.table.item(r, PRESET_COL_ADDR)
            di = self.table.item(r, PRESET_COL_DATA)
            data[str(r)] = [
                ai.text().strip() if ai else "00",
                di.text().strip() if di else "00",
            ]
        return data

    def _show_preset(self, name):
        rows = self._presets_db.get(name, _default_preset())
        self.table.blockSignals(True)
        for r in range(NUM_ROWS):
            row = rows.get(str(r), ["00", "00"])
            for c, val in [(PRESET_COL_ADDR, row[0]), (PRESET_COL_DATA, row[1])]:
                item = self.table.item(r, c)
                if item is None:
                    item = QTableWidgetItem()
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(r, c, item)
                item.setText(str(val))
        self.table.blockSignals(False)

    def _on_preset_activated(self, index):
        name = self.cb_preset.itemText(index)
        if name and name in self._presets_db:
            self._show_preset(name)

    def _save_preset(self):
        name = self.cb_preset.currentText().strip()
        exists = name and name in self._presets_db

        if exists:
            btn = QMessageBox.question(
                self, "保存预设",
                f"预设 '{name}' 已存在。\n"
                f"覆盖 '{name}' 请选 [Yes]\n"
                f"另存为新建请选 [No]\n"
                f"取消请选 [Cancel]",
                QMessageBox.StandardButton.Yes
                | QMessageBox.StandardButton.No
                | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Cancel,
            )
            if btn == QMessageBox.StandardButton.Cancel:
                return
            if btn == QMessageBox.StandardButton.No:
                name = ""

        if not name:
            name, ok = QInputDialog.getText(
                self, "新建预设", "预设名称:",
                text="" if exists else name,
            )
            if not ok or not name or not name.strip():
                return
            name = name.strip()

        self._presets_db[name] = self._gather_right_side()
        self._save_preset_db()

        self.cb_preset.blockSignals(True)
        if self.cb_preset.findText(name) == -1:
            self.cb_preset.addItem(name)
        self.cb_preset.setCurrentText(name)
        self.cb_preset.blockSignals(False)

        self.log_signal.emit(f"[INFO] 预设 '{name}' 已保存")

    def _rename_preset(self):
        old_name = self.cb_preset.currentText().strip()
        if not old_name or old_name not in self._presets_db:
            return
        new_name, ok = QInputDialog.getText(
            self, "重命名预设", "新名称:", text=old_name)
        if not ok:
            return
        new_name = new_name.strip()
        if not new_name or new_name == old_name:
            return
        if new_name in self._presets_db:
            QMessageBox.warning(self, "重命名失败",
                f"预设名称 '{new_name}' 已存在，请换一个名称。")
            return
        self._presets_db[new_name] = self._presets_db.pop(old_name)
        self._save_preset_db()

        self.cb_preset.blockSignals(True)
        idx = self.cb_preset.findText(old_name)
        if idx >= 0:
            self.cb_preset.setItemText(idx, new_name)
        self.cb_preset.setCurrentText(new_name)
        self.cb_preset.blockSignals(False)

        self.log_signal.emit(f"[INFO] 预设 '{old_name}' 已重命名为 '{new_name}'")

    def _delete_preset(self):
        name = self.cb_preset.currentText().strip()
        if not name or name not in self._presets_db:
            return
        r = QMessageBox.question(self, "删除预设",
            f"确认删除预设 '{name}' ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if r != QMessageBox.StandardButton.Yes:
            return
        del self._presets_db[name]
        self._save_preset_db()

        self.cb_preset.blockSignals(True)
        idx = self.cb_preset.findText(name)
        if idx >= 0:
            self.cb_preset.removeItem(idx)
        if self.cb_preset.count() > 0:
            self.cb_preset.setCurrentIndex(0)
        self.cb_preset.blockSignals(False)

        self.log_signal.emit(f"[INFO] 预设 '{name}' 已删除")

    def _fill_from_presets(self):
        name = self.cb_preset.currentText().strip()
        rows = self._presets_db.get(name, _default_preset())
        self.table.blockSignals(True)
        for r in range(NUM_ROWS):
            row = rows.get(str(r), ["00", "00"])
            for c, val in [(SEND_COL_ADDR, row[0]), (SEND_COL_DATA, row[1])]:
                item = self.table.item(r, c)
                if item is None:
                    item = QTableWidgetItem()
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(r, c, item)
                item.setText(str(val))
        self.table.blockSignals(False)
        self.log_signal.emit(f"[INFO] 已从预设 '{name}' 填至左侧")

    # ----------------------------------------------------------------
    # SPI 操作
    # ----------------------------------------------------------------

    def _send_config(self):
        if not self.link.is_open():
            QMessageBox.warning(self, "未连接", "串口未打开")
            return
        cpol = int(self.cb_cpol.currentText())
        cpha = int(self.cb_cpha.currentText())
        prescaler = self.cb_prescaler.currentData()
        dwidth = self.cb_dwidth.currentData()
        frame = build_spi_config(cpol, cpha, prescaler, dwidth)
        self.link.send(frame.to_bytes())
        self.log_signal.emit(
            f"[TX] SPI配置 CPOL={cpol} CPHA={cpha} "
            f"分频={self.cb_prescaler.currentText()} {dwidth}-bit"
        )

    def _send_regs(self):
        if not self.link.is_open():
            return
        regs: List[Tuple[int, int]] = []
        dwidth = self.cb_dwidth.currentData()
        for r in range(NUM_ROWS):
            if not self._row_checks[r].isChecked():
                continue
            ai = self.table.item(r, SEND_COL_ADDR)
            di = self.table.item(r, SEND_COL_DATA)
            if not ai or not di:
                continue
            at = ai.text().strip()
            dt = di.text().strip()
            if not at or not dt:
                continue
            try:
                regs.append((int(at, 16), int(dt, 16)))
            except ValueError:
                continue
        if not regs:
            return
        frame = build_spi_reg_writes(regs, data_bits=dwidth)
        if len(frame.to_bytes()) > 2048 + 8:
            return
        self.link.send(frame.to_bytes())
        self.log_signal.emit(f"[TX] SPI 写入 {len(regs)} 个寄存器, {dwidth}-bit")

    def _toggle_nonstop(self, state):
        if state == Qt.CheckState.Checked.value:
            self.sb_interval.setEnabled(False)
            self.sb_interval.setValue(1)
            if self.chk_burst.isChecked():
                self._burst_timer.stop()
                self._burst_timer.start(1)
                self.log_signal.emit("[INFO] SPI 不间断发送(1ms)")
        else:
            self.sb_interval.setEnabled(True)
            if self.chk_burst.isChecked():
                self._burst_timer.stop()
                self._burst_timer.start(self.sb_interval.value())

    def _toggle_burst(self, state):
        if state == Qt.CheckState.Checked.value:
            iv = 1 if self.chk_nonstop.isChecked() else self.sb_interval.value()
            self._burst_timer.start(iv)
            self.log_signal.emit(f"[INFO] SPI 持续发送启动({iv}ms)")
        else:
            self._burst_timer.stop()
            self.log_signal.emit("[INFO] SPI 持续发送已停止")

    def _on_header_clicked(self, idx):
        if idx == SEND_COL_CHECK:
            # Toggle: if all checked → uncheck all; else → check all
            all_checked = all(chk.isChecked() for chk in self._row_checks)
            for chk in self._row_checks:
                chk.setChecked(not all_checked)

    # ----------------------------------------------------------------
    # Excel 导入导出
    # ----------------------------------------------------------------
    _EXPORT_DIR = r"D:\test\STM32H723ZGT6\host\spi_contest"

    def _export_excel(self):
        if not self._presets_db:
            QMessageBox.information(self, "导出失败", "没有可导出的预设。")
            return
        os.makedirs(self._EXPORT_DIR, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(self._EXPORT_DIR, f"spi_presets_{ts}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet
        for name, rows in self._presets_db.items():
            ws = wb.create_sheet(title=name[:31])  # Excel sheet name max 31 chars
            ws.cell(row=1, column=1, value="行号")
            ws.cell(row=1, column=2, value="地址(Hex)")
            ws.cell(row=1, column=3, value="数据(Hex)")
            for r in range(NUM_ROWS):
                row = rows.get(str(r), ["00", "00"])
                ws.cell(row=r + 2, column=1, value=r)
                ws.cell(row=r + 2, column=2, value=row[0])
                ws.cell(row=r + 2, column=3, value=row[1])
        wb.save(path)
        self.log_signal.emit(f"[INFO] 预设已导出到: {path}")

    def _import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "导入 SPI 预设", self._EXPORT_DIR,
            "Excel 文件 (*.xlsx);;所有文件 (*)")
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"无法读取 Excel 文件:\n{e}")
            return

        imported = 0
        overwritten = []
        for ws in wb.worksheets:
            name = ws.title.strip()
            if not name:
                continue
            preset = {}
            for row_idx in range(2, 2 + NUM_ROWS):
                r = row_idx - 2
                addr_cell = ws.cell(row=row_idx, column=2).value
                data_cell = ws.cell(row=row_idx, column=3).value
                addr = str(addr_cell).strip() if addr_cell is not None else "00"
                data = str(data_cell).strip() if data_cell is not None else "00"
                preset[str(r)] = [addr, data]
            if name in self._presets_db:
                overwritten.append(name)
            self._presets_db[name] = preset
            imported += 1

        self._save_preset_db()

        # Refresh combo box
        self.cb_preset.blockSignals(True)
        self.cb_preset.clear()
        names = sorted(self._presets_db.keys())
        self.cb_preset.addItems(names)
        if names:
            self.cb_preset.setCurrentIndex(0)
        self.cb_preset.blockSignals(False)
        if names:
            self._show_preset(names[0])

        msg = f"已导入 {imported} 个预设"
        if overwritten:
            msg += f"\n已覆盖: {', '.join(overwritten)}"
        self.log_signal.emit(f"[INFO] {msg}")
        QMessageBox.information(self, "导入完成", msg)

    def on_frame(self, frame):
        from comm.protocol import CMD_ACK
        if frame.cmd == CMD_ACK:
            if len(frame.payload) < 2:
                return
            ok = "成功" if frame.payload[1] == 0 else "失败"
            self.log_signal.emit(f"[RX] 确认 cmd=0x{frame.payload[0]:02X} {ok}")
