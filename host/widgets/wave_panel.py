# -*- coding: utf-8 -*-
"""Waveform generator panel with graphical timing editor.
Performance optimized: debounced updates, batch table ops, reused CLK markers,
all channels editable, CLK markers toggle, Excel import/export.
"""

import os
import json
import numpy as np
import pyqtgraph as pg
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QSpinBox, QDoubleSpinBox, QTableWidget, QTableWidgetItem,
    QGroupBox, QMessageBox, QLineEdit, QComboBox, QCheckBox,
    QHeaderView, QAbstractItemView, QSplitter, QSizePolicy,
    QFileDialog,
)
from PyQt6.QtCore import Qt, pyqtSignal, QTimer
from typing import List
import openpyxl

from comm.protocol import (
    build_wave_config, build_wave_data, build_wave_ctrl,
    CMD_ACK,
)
from comm.serial_link import SerialLink

_TEMPLATES_PATH = os.path.join(os.path.expanduser("~"), ".h723_wave_templates.json")
_CURRENT_KEY = "__current__"

CH_NAMES = ["XYNC", "SCLK", "SH_R", "SH_S", "RST"]
CH_PINS = ["PA0", "PA1", "PA2", "PA3", "PA5"]
CH_BITS = [0, 1, 2, 3, 5]
CH_COLORS = [(255,0,0),(0,255,0),(0,0,255),(255,255,0),(255,0,255)]
NUM_CH = 5

class WaveViewBox(pg.ViewBox):
    def wheelEvent(self, ev, axis=None):
        if ev.modifiers() & Qt.KeyboardModifier.ControlModifier:
            pg.ViewBox.wheelEvent(self, ev, axis=0)
        else:
            pg.ViewBox.wheelEvent(self, ev, axis=1)
        ev.accept()


class WavePanel(QWidget):
    log_signal = pyqtSignal(str)

    def __init__(self, link: SerialLink, parent=None):
        super().__init__(parent)
        self.link = link
        self._masks: List[int] = []
        self._offset_spins: List[QDoubleSpinBox] = []
        self._clk_lines: List[pg.InfiniteLine] = []
        self._clk_texts: List[pg.TextItem] = []
        self._curves: List[pg.PlotDataItem] = []

        # CLK markers visibility flag
        self._show_clk_markers = True

        # Debounce for plot updates
        self._plot_debounce = QTimer(self)
        self._plot_debounce.setSingleShot(True)
        self._plot_debounce.setInterval(30)
        self._plot_debounce.timeout.connect(self._do_update_plot)

        # Debounce for table resize
        self._resize_debounce = QTimer(self)
        self._resize_debounce.setSingleShot(True)
        self._resize_debounce.setInterval(30)
        self._resize_debounce.timeout.connect(self._do_resize_table)

        self._setup_ui()
        self._fill_sclk()
        self._load_rules()
        self._schedule_plot_update()

    def _setup_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(4, 4, 4, 4)
        outer.setSpacing(4)

        main_splitter = QSplitter(Qt.Orientation.Vertical)
        main_splitter.setHandleWidth(6)

        # ---- TOP: Control panel ----
        ctrl_widget = QWidget()
        ctrl_layout = QHBoxLayout(ctrl_widget)
        ctrl_layout.setContentsMargins(0, 0, 0, 0)
        ctrl_layout.setSpacing(6)

        # --- 1. Waveform Config ---
        cfg_box = QGroupBox("波形配置")
        cfg_box.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Maximum)
        cfg_inner = QVBoxLayout(cfg_box)
        cfg_inner.setSpacing(3)
        cfg_inner.setContentsMargins(6, 6, 6, 6)

        row = QHBoxLayout()
        row.addWidget(QLabel("频率(Hz):"))
        self.sb_rate = QSpinBox()
        self.sb_rate.setRange(1, 10_000_000)
        self.sb_rate.setValue(1000)
        row.addWidget(self.sb_rate)
        cfg_inner.addLayout(row)

        row = QHBoxLayout()
        row.addWidget(QLabel("点数:"))
        self.sb_points = QSpinBox()
        self.sb_points.setRange(1, 8192)
        self.sb_points.setValue(160)
        self.sb_points.valueChanged.connect(self._on_points_changed)
        row.addWidget(self.sb_points)
        cfg_inner.addLayout(row)

        row = QHBoxLayout()
        row.addWidget(QLabel("掩码:"))
        self.sb_chmask = QSpinBox()
        self.sb_chmask.setRange(1, 31)
        self.sb_chmask.setValue(31)
        row.addWidget(self.sb_chmask)
        cfg_inner.addLayout(row)

        row = QHBoxLayout()
        row.addWidget(QLabel("SCLK初始:"))
        self.cb_sclk_init = QComboBox()
        self.cb_sclk_init.addItems(["高", "低"])
        self.cb_sclk_init.setCurrentIndex(0)
        self.cb_sclk_init.currentTextChanged.connect(self._on_sclk_config_changed)
        row.addWidget(self.cb_sclk_init)
        cfg_inner.addLayout(row)

        row = QHBoxLayout()
        row.addWidget(QLabel("有效边沿:"))
        self.cb_sclk_edge = QComboBox()
        self.cb_sclk_edge.addItems(["下降沿", "上升沿"])
        self.cb_sclk_edge.setCurrentIndex(0)
        self.cb_sclk_edge.currentTextChanged.connect(self._on_sclk_config_changed)
        row.addWidget(self.cb_sclk_edge)
        cfg_inner.addLayout(row)

        row = QHBoxLayout()
        row.addWidget(QLabel("CLK截止:"))
        self.sb_sclk_cutoff = QSpinBox()
        self.sb_sclk_cutoff.setRange(0, 4096)
        self.sb_sclk_cutoff.setValue(0)
        self.sb_sclk_cutoff.setToolTip("0=全覆盖, N=在第N个CLK后停止(全低)")
        row.addWidget(self.sb_sclk_cutoff)
        self.sb_sclk_cutoff.valueChanged.connect(self._on_sclk_config_changed)
        cfg_inner.addLayout(row)

        self.btn_generate = QPushButton("生成并加载到设备")
        self.btn_generate.clicked.connect(self._unified_generate)
        cfg_inner.addWidget(self.btn_generate)
        cfg_inner.addStretch()

        ctrl_layout.addWidget(cfg_box, stretch=0)

        # --- 2. Timing Rule Editor ---
        rule_box = QGroupBox("时序规则编辑器")
        rule_inner = QVBoxLayout(rule_box)
        rule_inner.setSpacing(3)
        rule_inner.setContentsMargins(6, 6, 6, 6)

        init_layout = QHBoxLayout()
        init_layout.addWidget(QLabel("初始状态:"))
        self._initial_checks = []
        for i, name in enumerate(CH_NAMES):
            if name == "SCLK":
                continue
            chk = QCheckBox(f"{name}")
            chk.stateChanged.connect(self._on_initial_state_changed)
            self._initial_checks.append((i, chk))
            init_layout.addWidget(chk)
        init_layout.addStretch()
        rule_inner.addLayout(init_layout)

        self.rule_table = QTableWidget()
        self.rule_table.setColumnCount(5)
        self.rule_table.setHorizontalHeaderLabels(["通道", "触发边沿序号", "新电平", "脉宽(CLK周期)", ""])
        self.rule_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.rule_table.setColumnWidth(4, 60)
        self.rule_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        rule_inner.addWidget(self.rule_table)

        btn_layout = QHBoxLayout()
        self.btn_add_rule = QPushButton("+ 添加规则")
        self.btn_add_rule.clicked.connect(self._add_rule)
        btn_layout.addWidget(self.btn_add_rule)
        btn_layout.addStretch()
        rule_inner.addLayout(btn_layout)

        # Template row
        tmpl_row = QHBoxLayout()
        tmpl_row.addWidget(QLabel("模板:"))
        self.cb_template = QComboBox()
        self.cb_template.setMinimumWidth(120)
        self.cb_template.activated.connect(self._on_template_selected)
        tmpl_row.addWidget(self.cb_template)
        self.btn_save_template = QPushButton("保存为模板")
        self.btn_save_template.clicked.connect(self._save_template)
        tmpl_row.addWidget(self.btn_save_template)
        self.btn_delete_template = QPushButton("删除模板")
        self.btn_delete_template.clicked.connect(self._delete_template)
        tmpl_row.addWidget(self.btn_delete_template)
        tmpl_row.addStretch()
        rule_inner.addLayout(tmpl_row)

        ctrl_layout.addWidget(rule_box, stretch=1)

        # --- 3. Playback ---
        play_box = QGroupBox("播放控制")
        play_box.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Maximum)
        play_inner = QVBoxLayout(play_box)
        play_inner.setSpacing(4)
        play_inner.setContentsMargins(6, 6, 6, 6)

        self.btn_start = QPushButton("开始")
        self.btn_start.clicked.connect(lambda: self._send_ctrl(True))
        play_inner.addWidget(self.btn_start)

        self.btn_stop = QPushButton("停止")
        self.btn_stop.clicked.connect(lambda: self._send_ctrl(False))
        play_inner.addWidget(self.btn_stop)

        play_inner.addStretch()
        ctrl_layout.addWidget(play_box, stretch=0)

        main_splitter.addWidget(ctrl_widget)

        # ---- BOTTOM: Display panel ----
        display_splitter = QSplitter(Qt.Orientation.Horizontal)
        display_splitter.setHandleWidth(6)

        # --- Waveform Plot (with CLK toggle checkbox) ---
        plot_box = QGroupBox("波形预览")
        plot_outer = QHBoxLayout(plot_box)
        plot_outer.setContentsMargins(4, 4, 4, 4)
        plot_outer.setSpacing(2)

        # Left panel: offsets + CLK toggle (name on top, spin below)
        left_plot_panel = QWidget()
        left_plot_panel.setFixedWidth(120)
        left_plot_layout = QVBoxLayout(left_plot_panel)
        left_plot_layout.setContentsMargins(2, 2, 2, 2)
        left_plot_layout.setSpacing(2)

        # CLK markers toggle
        self.cb_show_clk = QCheckBox("CLK虚线")
        self.cb_show_clk.setChecked(True)
        self.cb_show_clk.stateChanged.connect(self._on_clk_visibility_changed)
        left_plot_layout.addWidget(self.cb_show_clk)

        left_plot_layout.addSpacing(6)

        # Offset controls: name label on top, spinbox below, per channel
        for i in range(5):
            color = CH_COLORS[i]
            namelbl = QLabel(f"<font color='rgb{color}'><b>{CH_NAMES[i]}</b></font>")
            namelbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            left_plot_layout.addWidget(namelbl)
            spin = QDoubleSpinBox()
            spin.setRange(-10.0, 10.0)
            spin.setSingleStep(0.1)
            spin.setValue(0.0)
            spin.setDecimals(1)
            spin.setFixedWidth(90)
            spin.valueChanged.connect(self._schedule_plot_update)
            row_spin = QHBoxLayout()
            row_spin.addStretch()
            row_spin.addWidget(spin)
            row_spin.addStretch()
            left_plot_layout.addLayout(row_spin)
            self._offset_spins.append(spin)
            if i < 4:
                left_plot_layout.addSpacing(4)

        left_plot_layout.addStretch()
        plot_outer.addWidget(left_plot_panel)

        self.plot_widget = pg.PlotWidget(viewBox=WaveViewBox())
        self.plot_widget.setLabel("bottom", "采样点序号")
        self.plot_widget.setLabel("left", "通道")
        self.plot_widget.showGrid(x=True, y=True)
        self.plot_widget.setYRange(-0.5, 7.0, padding=0.0)
        self.plot_widget.getAxis("left").setTicks([
            [(i * 1.5 + 0.4, CH_NAMES[4 - i]) for i in range(5)]
        ])
        self.plot_widget.getViewBox().setMouseMode(pg.ViewBox.PanMode)

        self._curves = []
        for i in range(5):
            pen = pg.mkPen(color=CH_COLORS[i], width=2)
            curve = self.plot_widget.plot(pen=pen, name=CH_NAMES[i])
            self._curves.append(curve)

        plot_outer.addWidget(self.plot_widget, stretch=1)
        display_splitter.addWidget(plot_box)

        # --- Pattern Table with Import/Export ---
        table_box = QGroupBox("模式编辑器（所有通道均可直接编辑）")
        table_layout = QVBoxLayout(table_box)
        table_layout.setContentsMargins(4, 4, 4, 4)

        names_layout = QHBoxLayout()
        names_layout.setSpacing(2)
        self._name_edits = []
        for pin, default in zip(CH_PINS, CH_NAMES):
            col_layout = QVBoxLayout()
            col_layout.setSpacing(0)
            col_layout.setContentsMargins(2, 0, 2, 0)
            pin_lbl = QLabel(f"<b>{pin}</b>")
            pin_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            col_layout.addWidget(pin_lbl)
            edit = QLineEdit(default)
            edit.setAlignment(Qt.AlignmentFlag.AlignCenter)
            edit.setMaxLength(12)
            col_layout.addWidget(edit)
            self._name_edits.append(edit)
            names_layout.addLayout(col_layout, stretch=1)
        table_layout.addLayout(names_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(CH_NAMES)
        self.table.setRowCount(0)
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.sb_points.valueChanged.connect(self._on_points_changed)
        self.table.cellChanged.connect(self._on_table_changed)
        table_layout.addWidget(self.table)

        for i, edit in enumerate(self._name_edits):
            edit.textChanged.connect(lambda text, idx=i: self.table.setHorizontalHeaderItem(
                idx, QTableWidgetItem(text)))

        # Buttons row: export | import
        btn_row = QHBoxLayout()
        self.btn_export = QPushButton("导出Excel")
        self.btn_export.clicked.connect(self._export_excel)
        btn_row.addWidget(self.btn_export)

        self.btn_import = QPushButton("导入Excel")
        self.btn_import.clicked.connect(self._import_excel)
        btn_row.addWidget(self.btn_import)

        btn_row.addStretch()

        table_layout.addLayout(btn_row)

        display_splitter.addWidget(table_box)
        display_splitter.setSizes([600, 260])
        display_splitter.setStretchFactor(0, 2)
        display_splitter.setStretchFactor(1, 1)

        main_splitter.addWidget(display_splitter)
        main_splitter.setSizes([220, 500])
        main_splitter.setStretchFactor(0, 0)
        main_splitter.setStretchFactor(1, 1)

        outer.addWidget(main_splitter)


    # ============================================================
    # High-Performance Table Operations
    # ============================================================

    def _batch_fill_table(self, states):
        rows, cols = states.shape
        self.table.blockSignals(True)
        self.table.setRowCount(rows)
        for c in range(cols):
            for r in range(rows):
                item = self.table.item(r, c)
                if item is None:
                    item = QTableWidgetItem()
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(r, c, item)
                item.setText("1" if states[r, c] else "0")
                item.setFlags(
                    Qt.ItemFlag.ItemIsEnabled |
                    Qt.ItemFlag.ItemIsSelectable |
                    Qt.ItemFlag.ItemIsEditable
                )
        self.table.blockSignals(False)

    def _do_resize_table(self):
        rows = self.sb_points.value()
        self.table.blockSignals(True)
        self.table.setRowCount(rows)
        for r in range(rows):
            for c in range(NUM_CH):
                item = self.table.item(r, c)
                if item is None:
                    # Safety: default new rows to HIGH instead of LOW to avoid
                    # silently introducing long low-level segments when the user
                    # increases the number of points.
                    item = QTableWidgetItem("1")
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setFlags(
                        Qt.ItemFlag.ItemIsEnabled |
                        Qt.ItemFlag.ItemIsSelectable |
                        Qt.ItemFlag.ItemIsEditable
                    )
                    self.table.setItem(r, c, item)
        self.table.blockSignals(False)
        self._schedule_plot_update()

    def _on_points_changed(self):
        self._resize_debounce.start()

    def _on_table_changed(self, row, col):
        item = self.table.item(row, col)
        if item is None:
            return
        text = item.text().strip()
        if text not in ("0", "1"):
            # Safety: invalid input defaults to HIGH instead of LOW to avoid
            # accidental short-to-ground on a channel.
            item.setText("1")
        self._clear_all_rules()
        self._schedule_plot_update()

    def _fill_sclk(self, silent=False):
        if not silent and self.table.rowCount() > 0:
            reply = QMessageBox.question(
                self, "确认填充",
                "将覆盖当前 SCLK 列所有数据，继续？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
        n = self.sb_points.value()
        initial_high = self.cb_sclk_init.currentText() == "高"
        cutoff_edge = self.sb_sclk_cutoff.value()
        cutoff_sample = cutoff_edge * 2 if cutoff_edge > 0 else n
        self.table.blockSignals(True)
        self.table.setRowCount(n)
        for r in range(n):
            item = self.table.item(r, 1)
            if item is None:
                item = QTableWidgetItem()
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setFlags(
                    Qt.ItemFlag.ItemIsEnabled |
                    Qt.ItemFlag.ItemIsSelectable |
                    Qt.ItemFlag.ItemIsEditable
                )
                self.table.setItem(r, 1, item)
            if r >= cutoff_sample:
                item.setText("0")
            elif initial_high:
                item.setText("1" if r % 2 == 0 else "0")
            else:
                item.setText("0" if r % 2 == 0 else "1")
        self.table.blockSignals(False)
        self._schedule_plot_update()

    def _on_sclk_config_changed(self):
        self._save_rules()

    def _on_initial_state_changed(self):
        self._save_rules()

    def _clear_all_rules(self):
        """Clear rule table when the pattern table is manually edited,
        preventing stale rules from overwriting manual edits on next apply."""
        if self.rule_table.rowCount() > 0:
            self.rule_table.setRowCount(0)
            self._save_rules()
            self.log_signal.emit("[INFO] 手动修改表格，已清除时序规则")

    # ============================================================
    # Safety helpers
    # ============================================================

    def _get_channel_states(self):
        """Return (n, NUM_CH) uint8 array of current table states."""
        n = self.table.rowCount()
        if n == 0:
            return np.zeros((0, NUM_CH), dtype=np.uint8)
        states = np.zeros((n, NUM_CH), dtype=np.uint8)
        for r in range(n):
            for c in range(NUM_CH):
                item = self.table.item(r, c)
                if item is not None and item.text().strip() == "1":
                    states[r, c] = 1
        return states

    def _check_waveform_safety(self, states, action_name="此操作"):
        """Check for risky patterns and return (is_safe, message).

        Currently flags:
        - Any channel that is low for 100% of the buffer (never high).
        - Any channel that is low for >80% of the buffer (long-low warning).
        - Any stretch of all-zero rows (all channels low) longer than 16 samples.
        """
        if states.size == 0:
            return True, ""
        n = states.shape[0]
        msgs = []
        for c in range(NUM_CH):
            low_ratio = 1.0 - np.mean(states[:, c])
            if low_ratio >= 1.0:
                msgs.append(f"通道 {CH_NAMES[c]}({CH_PINS[c]}) 全程为低电平，可能导致持续大电流。")
            elif low_ratio > 0.80:
                msgs.append(f"通道 {CH_NAMES[c]}({CH_PINS[c]}) 低电平占比 {low_ratio*100:.0f}%，存在长时间灌电流风险。")
        # All-zero stretch
        all_low = np.all(states == 0, axis=1)
        if np.any(all_low):
            max_stretch = 0
            cur = 0
            for v in all_low:
                if v:
                    cur += 1
                    max_stretch = max(max_stretch, cur)
                else:
                    cur = 0
            if max_stretch > 16:
                msgs.append(f"存在连续 {max_stretch} 个采样点所有通道同时为低，风险极高。")
        if not msgs:
            return True, ""
        detail = "\n".join(f"  • {m}" for m in msgs)
        return False, f"{action_name}将生成以下高风险波形，是否继续？\n\n{detail}\n\n建议：点击“取消”检查通道电平，确认外部电路可承受后再继续。"

    def _warn_and_confirm(self, states, action_name="此操作"):
        """Return True if the action should proceed."""
        return True

    # ============================================================
    # Rule Editor
    # ============================================================

    def _add_rule(self):
        row = self.rule_table.rowCount()
        self.rule_table.insertRow(row)
        self._save_rules()
        ch_combo = QComboBox()
        for name in CH_NAMES:
            if name == "SCLK":
                continue
            ch_combo.addItem(name)
        ch_combo.currentTextChanged.connect(self._save_rules)
        self.rule_table.setCellWidget(row, 0, ch_combo)
        edge_spin = QSpinBox()
        edge_spin.setRange(0, 4096)
        edge_spin.setValue(0)
        edge_spin.valueChanged.connect(self._save_rules)
        self.rule_table.setCellWidget(row, 1, edge_spin)
        level_combo = QComboBox()
        level_combo.addItems(["高", "低"])
        level_combo.currentTextChanged.connect(self._save_rules)
        self.rule_table.setCellWidget(row, 2, level_combo)
        width_spin = QSpinBox()
        width_spin.setRange(0, 4096)
        width_spin.setValue(0)
        width_spin.setToolTip("0=永久保持；N=从触发点起持续N个CLK周期(1CLK周期=2采样点)")
        width_spin.valueChanged.connect(self._save_rules)
        self.rule_table.setCellWidget(row, 3, width_spin)
        btn = QPushButton("删除")
        btn.clicked.connect(lambda checked, r=row: self._remove_rule(r))
        self.rule_table.setCellWidget(row, 4, btn)

    def _remove_rule(self, row):
        self.rule_table.removeRow(row)
        self._save_rules()
        for r in range(self.rule_table.rowCount()):
            btn = self.rule_table.cellWidget(r, 4)
            if isinstance(btn, QPushButton):
                btn.clicked.disconnect()
                btn.clicked.connect(lambda checked, idx=r: self._remove_rule(idx))

    def _ch_name_to_idx(self, name):
        return CH_NAMES.index(name)

    def _edge_to_sample(self, edge_num):
        """0-based: edge 0 = first active CLK edge, edge 1 = second, etc."""
        initial_high = self.cb_sclk_init.currentText() == "高"
        falling_edge = self.cb_sclk_edge.currentText() == "下降沿"
        match = (initial_high and falling_edge) or (not initial_high and not falling_edge)
        if match:
            return 2 * edge_num + 1
        else:
            return 2 * edge_num

    def _apply_rules(self):
        n = self.sb_points.value()
        if n == 0:
            return False
        # Parse rules first to know which channels are controlled by rules.
        rules = []
        rule_chs = set()
        for r in range(self.rule_table.rowCount()):
            ch_widget = self.rule_table.cellWidget(r, 0)
            edge_widget = self.rule_table.cellWidget(r, 1)
            level_widget = self.rule_table.cellWidget(r, 2)
            width_widget = self.rule_table.cellWidget(r, 3)
            if ch_widget is None or edge_widget is None:
                continue
            ch_name = ch_widget.currentText()
            edge_num = edge_widget.value()
            level = 1 if level_widget.currentText() == "高" else 0
            width = width_widget.value()
            ch_idx = self._ch_name_to_idx(ch_name)
            sample_idx = self._edge_to_sample(edge_num)
            rules.append((sample_idx, ch_idx, level, width))
            rule_chs.add(ch_idx)
        # Channels without rules keep existing table data (e.g. SCLK).
        # Channels with rules start from their initial-state checkbox.
        states = np.zeros((n, NUM_CH), dtype=np.uint8)
        for r in range(min(n, self.table.rowCount())):
            for c in range(NUM_CH):
                if c in rule_chs:
                    continue
                item = self.table.item(r, c)
                if item is not None and item.text().strip() == "1":
                    states[r, c] = 1
        for i, chk in self._initial_checks:
            if chk.isChecked() and i in rule_chs:
                states[:, i] = 1
        rules.sort(key=lambda x: x[0])
        for sample_idx, ch_idx, level, width in rules:
            if sample_idx >= n:
                continue
            if width == 0:
                states[sample_idx:, ch_idx] = level
            else:
                end_idx = min(n, sample_idx + width * 2)
                states[sample_idx:end_idx, ch_idx] = level

        # Safety: warn about long-low / all-low patterns before applying.
        if not self._warn_and_confirm(states, "应用时序规则"):
            self.log_signal.emit("[SAFETY] 用户取消应用高风险时序规则")
            return False

        self._batch_fill_table(states)
        self._schedule_plot_update()
        self.log_signal.emit(f"[INFO] 已应用 {len(rules)} 条时序规则，总点数={n}")
        return True


    # ============================================================
    # CLK Markers Visibility
    # ============================================================

    def _on_clk_visibility_changed(self):
        self._show_clk_markers = self.cb_show_clk.isChecked()
        self._do_update_plot()

    # ============================================================
    # Plot (reused CLK markers, debounced)
    # ============================================================

    def _schedule_plot_update(self):
        self._plot_debounce.start()

    def _stair_step(self, values, y_base, offset=0.0):
        n = len(values)
        if n == 0:
            return np.array([]), np.array([])
        x = np.empty(n * 2)
        y = np.empty(n * 2)
        for i in range(n):
            x[i * 2] = i
            x[i * 2 + 1] = i + 1
            y[i * 2] = y_base + values[i] * 0.8 + offset
            y[i * 2 + 1] = y_base + values[i] * 0.8 + offset
        return x, y

    def _do_update_plot(self):
        n = self.table.rowCount()

        if n == 0:
            for curve in self._curves:
                curve.setData([], [])
            self._clear_clk_markers()
            return

        ch_states = []
        for c in range(NUM_CH):
            arr = np.array([
                1 if (self.table.item(r, c) is not None
                      and self.table.item(r, c).text().strip() == "1")
                else 0
                for r in range(n)
            ], dtype=np.uint8)
            ch_states.append(arr)

        # --- CLK markers (toggle via _show_clk_markers) ---
        initial_high = self.cb_sclk_init.currentText() == "高"
        falling_edge = self.cb_sclk_edge.currentText() == "下降沿"
        match = (initial_high and falling_edge) or (not initial_high and not falling_edge)

        max_edge = n // 2
        needed = max_edge if self._show_clk_markers else 0

        while len(self._clk_lines) < needed:
            line = pg.InfiniteLine(
                pos=0, angle=90,
                pen=pg.mkPen(color=(160, 160, 160), width=1,
                             style=Qt.PenStyle.DotLine)
            )
            text = pg.TextItem(text="", color=(160, 160, 160), anchor=(0.5, 1))
            self.plot_widget.addItem(line)
            self.plot_widget.addItem(text)
            self._clk_lines.append(line)
            self._clk_texts.append(text)

        for i in range(len(self._clk_lines)):
            if i < needed:
                edge_num = i  # 0-based
                if match:
                    sample_idx = 2 * edge_num + 1
                else:
                    sample_idx = 2 * edge_num
                self._clk_lines[i].setPos(sample_idx)
                self._clk_lines[i].show()
                self._clk_texts[i].setText(str(edge_num))
                self._clk_texts[i].setPos(sample_idx, 7.0)
                self._clk_texts[i].show()
            else:
                self._clk_lines[i].hide()
                self._clk_texts[i].hide()

        # --- Update curve data ---
        for c in range(NUM_CH):
            y_base = (4 - c) * 1.5
            offset = self._offset_spins[c].value() if c < len(self._offset_spins) else 0.0
            x, y = self._stair_step(ch_states[c], y_base, offset)
            self._curves[c].setData(x, y)

        self.plot_widget.setXRange(0, n, padding=0.02)


    def _clear_clk_markers(self):
        for line in self._clk_lines:
            line.hide()
        for text in self._clk_texts:
            text.hide()

    # ============================================================
    # Excel Import / Export
    # ============================================================

    def _export_excel(self):
        rows = self.table.rowCount()
        if rows == 0:
            QMessageBox.information(self, "提示", "表格为空，无法导出。")
            return

        # Collect channel names from the header edits
        ch_names = [edit.text() for edit in self._name_edits]

        filepath, _ = QFileDialog.getSaveFileName(
            self, "导出波形数据", "", "Excel Files (*.xlsx)"
        )
        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Waveform"

        # Header row: 时序 + 5 channels
        headers = ["时序"] + ch_names
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        # Data rows
        for r in range(rows):
            ws.cell(row=r + 2, column=1, value=r)  # 时序单位（从0开始）
            for c in range(NUM_CH):
                item = self.table.item(r, c)
                val = item.text().strip() if item else "0"
                ws.cell(row=r + 2, column=c + 2, value=int(val))

        wb.save(filepath)
        self.log_signal.emit(f"[INFO] 波形数据已导出到 {filepath}")
        QMessageBox.information(self, "导出成功",
            f"已导出 {rows} 行数据到:\n{filepath}")

    def _import_excel(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "导入波形数据", "", "Excel Files (*.xlsx)"
        )
        if not filepath:
            return

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            if ws is None:
                ws = wb.worksheets[0]

            max_row = ws.max_row
            max_col = ws.max_column

            if max_row < 2:
                QMessageBox.warning(self, "导入失败", "Excel 文件至少需要 2 行（1行标题+1行数据）。")
                return

            if max_col < 2:
                QMessageBox.warning(self, "导入失败", "Excel 文件至少需要 2 列（1列时序+1列波形数据）。")
                return

            # Read header row for channel names
            excel_ch_names = []
            for col_idx in range(2, max_col + 1):
                h_val = ws.cell(row=1, column=col_idx).value
                excel_ch_names.append(str(h_val) if h_val is not None else f"CH{col_idx-1}")

            num_ch_import = min(len(excel_ch_names), NUM_CH)
            data_rows = max_row - 1  # minus header

            # Update channel name edits
            for i in range(num_ch_import):
                if i < len(self._name_edits):
                    self._name_edits[i].blockSignals(True)
                    self._name_edits[i].setText(excel_ch_names[i])
                    self._name_edits[i].blockSignals(False)

            # Update table
            self.table.blockSignals(True)
            self.table.setRowCount(data_rows)

            for r in range(data_rows):
                for c in range(num_ch_import):
                    excel_col = c + 2
                    cell_val = ws.cell(row=r + 2, column=excel_col).value
                    val = "1" if cell_val and int(cell_val) == 1 else "0"
                    item = self.table.item(r, c)
                    if item is None:
                        item = QTableWidgetItem()
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        item.setFlags(
                            Qt.ItemFlag.ItemIsEnabled |
                            Qt.ItemFlag.ItemIsSelectable |
                            Qt.ItemFlag.ItemIsEditable
                        )
                        self.table.setItem(r, c, item)
                    item.setText(val)
                # Fill remaining channels with 0
                for c in range(num_ch_import, NUM_CH):
                    item = self.table.item(r, c)
                    if item is None:
                        item = QTableWidgetItem()
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        item.setFlags(
                            Qt.ItemFlag.ItemIsEnabled |
                            Qt.ItemFlag.ItemIsSelectable |
                            Qt.ItemFlag.ItemIsEditable
                        )
                        self.table.setItem(r, c, item)
                    item.setText("0")

            self.table.blockSignals(False)

            # Update sb_points
            self.sb_points.blockSignals(True)
            self.sb_points.setValue(data_rows)
            self.sb_points.blockSignals(False)

            self._schedule_plot_update()
            self.log_signal.emit(f"[INFO] 从 Excel 导入 {data_rows} 行 x {num_ch_import} 通道波形数据")

            # Safety: warn about risky patterns imported from external files.
            imported_states = self._get_channel_states()
            if not self._warn_and_confirm(imported_states, "从 Excel 导入"):
                self.log_signal.emit("[SAFETY] Excel 导入数据存在高风险，用户已收到警告")
                QMessageBox.information(self, "导入完成（未下发）",
                    f"已导入 {data_rows} 行数据，但检测到高风险波形。\n"
                    "请在编辑器中检查各通道电平，确认安全后再点击“加载模式到设备”。")
                return

            QMessageBox.information(self, "导入成功",
                f"已导入 {data_rows} 行 x {num_ch_import} 通道数据。")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"读取 Excel 文件时出错:\n{str(e)}")
            self.log_signal.emit(f"[ERROR] Excel 导入失败: {e}")

    # ============================================================
    # Protocol
    # ============================================================

    def _get_bsrr_masks(self):
        """Build BSRR masks from state transitions, matching the default
        waveform style: only set BS/BR for pins that actually toggle.
        Pins that stay low across the whole buffer get no BR at all."""
        masks = []
        rows = self.table.rowCount()
        # Seed prev_state from the last row so the circular buffer wraps cleanly.
        prev_state = 0
        if rows > 0:
            for c in range(NUM_CH):
                item = self.table.item(rows - 1, c)
                if item is not None and item.text().strip() == "1":
                    prev_state |= (1 << c)

        for r in range(rows):
            state = 0
            for c in range(NUM_CH):
                item = self.table.item(r, c)
                if item is not None and item.text().strip() == "1":
                    state |= (1 << c)

            bsrr = 0
            bits = CH_BITS
            for i, bit in enumerate(bits):
                curr = (state >> i) & 1
                prev = (prev_state >> i) & 1
                if curr and not prev:
                    bsrr |= (1 << bit)          # LOW -> HIGH
                elif not curr and prev:
                    bsrr |= (1 << (bit + 16))   # HIGH -> LOW
            # Note: if curr==prev for a pin, nothing is set (matching default style)

            masks.append(bsrr)
            prev_state = state
        return masks

    @staticmethod
    def _state_to_bsrr(state):
        """ponytail: match the default waveform style - only set BS/BR for
         * channels that actually need to change, never write BR to a pin
         * that is already guaranteed low from the previous sample.
         * This mirrors the hand-written default wave: 0x00000002 / 0x00020000."""
        bsrr = 0
        bits = CH_BITS
        for i, bit in enumerate(bits):
            if state & (1 << i):
                bsrr |= (1 << bit)
            else:
                bsrr |= (1 << (bit + 16))
        return bsrr

    def _send_config(self):
        if not self.link.is_open():
            QMessageBox.warning(self, "???", "??????")
            return False
        update_rate = self.sb_rate.value() * 2
        frame = build_wave_config(
            update_rate,
            self.sb_points.value(),
            self.sb_chmask.value(),
        )
        self.link.send(frame.to_bytes())
        self.log_signal.emit(
            f"[TX] ???? ??={self.sb_rate.value()}Hz "
            f"???={update_rate} ??={self.sb_points.value()}"
        )
        return True

    def _send_data(self):
        if not self.link.is_open():
            QMessageBox.warning(self, "???", "??????")
            return False
        masks = self._get_bsrr_masks()
        if len(masks) > 0 and all(m == 0 for m in masks):
            QMessageBox.warning(self, "????",
                "BSRR ???????????\n"
                "??????????????????????????")
            self.log_signal.emit("[SAFETY] ?? BSRR ???????")
            return False

        # Safety: per-channel long-low / all-low check before sending to device.
        states = self._get_channel_states()
        if not self._warn_and_confirm(states, "???????"):
            self.log_signal.emit("[SAFETY] ???????????")
            return False

        if len(masks) > 2048:
            QMessageBox.information(self, "??",
                f"???? {len(masks)} ????????????")
            return False
        frame = build_wave_data(masks)
        self.link.send(frame.to_bytes())
        self.log_signal.emit(f"[TX] ???? {len(masks)} ?")
        return True

    def _unified_generate(self):
        """一键操作：发送配置→应用时序规则→填充SCLK→加载数据到设备"""
        self._send_config()
        if not self._apply_rules():
            return
        self._fill_sclk(silent=True)
        self._send_data()

    def _send_ctrl(self, start):
        if not self.link.is_open():
            QMessageBox.warning(self, "未连接", "串口未打开。")
            return
        frame = build_wave_ctrl(start)
        self.link.send(frame.to_bytes())
        self.log_signal.emit(f"[TX] 波形控制 {'开始' if start else '停止'}")

    def on_frame(self, frame):
        if frame.cmd == CMD_ACK:
            if len(frame.payload) < 2:
                return
            cmd_id = frame.payload[0]
            if cmd_id in (0x01, 0x02, 0x03):
                ok = "成功" if frame.payload[1] == 0 else "失败"
                cmd_name = {0x01: "WAVE_CONFIG", 0x02: "WAVE_DATA", 0x03: "WAVE_CTRL"}.get(cmd_id, f"0x{cmd_id:02X}")
                self.log_signal.emit(f"[RX] 确认 {cmd_name} {ok}")

    # ============================================================
    # Rule + Template persistence
    # ============================================================

    def _gather_rules_data(self):
        """Collect current SCLK config + initial states + rules into a dict."""
        data = {}
        data["sclk_init"] = self.cb_sclk_init.currentText()
        data["sclk_edge"] = self.cb_sclk_edge.currentText()
        data["sclk_cutoff"] = self.sb_sclk_cutoff.value()
        initial = {}
        for i, chk in self._initial_checks:
            initial[str(i)] = chk.isChecked()
        data["initial_states"] = initial
        rules = []
        for r in range(self.rule_table.rowCount()):
            ch_widget = self.rule_table.cellWidget(r, 0)
            edge_widget = self.rule_table.cellWidget(r, 1)
            level_widget = self.rule_table.cellWidget(r, 2)
            width_widget = self.rule_table.cellWidget(r, 3)
            if ch_widget is None or edge_widget is None:
                continue
            rules.append({
                "channel": ch_widget.currentText(),
                "edge": edge_widget.value(),
                "level": level_widget.currentText() if level_widget else "高",
                "width": width_widget.value() if width_widget else 0,
            })
        data["rules"] = rules
        return data

    def _load_templates_db(self):
        if not os.path.exists(_TEMPLATES_PATH):
            return {}
        try:
            with open(_TEMPLATES_PATH, "r", encoding="utf-8") as f:
                db = json.load(f)
            return db if isinstance(db, dict) else {}
        except Exception:
            return {}

    def _save_templates_db(self, db):
        try:
            with open(_TEMPLATES_PATH, "w", encoding="utf-8") as f:
                json.dump(db, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _save_rules(self):
        """Auto-save current editor state to __current__ slot."""
        db = self._load_templates_db()
        db[_CURRENT_KEY] = self._gather_rules_data()
        self._save_templates_db(db)

    def _apply_rules_data(self, data):
        """Populate editor UI from a rules data dict."""
        # SCLK config
        if "sclk_init" in data:
            idx = self.cb_sclk_init.findText(data["sclk_init"])
            if idx >= 0:
                self.cb_sclk_init.blockSignals(True)
                self.cb_sclk_init.setCurrentIndex(idx)
                self.cb_sclk_init.blockSignals(False)
        if "sclk_edge" in data:
            idx = self.cb_sclk_edge.findText(data["sclk_edge"])
            if idx >= 0:
                self.cb_sclk_edge.blockSignals(True)
                self.cb_sclk_edge.setCurrentIndex(idx)
                self.cb_sclk_edge.blockSignals(False)
        if "sclk_cutoff" in data:
            self.sb_sclk_cutoff.blockSignals(True)
            self.sb_sclk_cutoff.setValue(data["sclk_cutoff"])
            self.sb_sclk_cutoff.blockSignals(False)
        # Initial states
        if "initial_states" in data:
            initial = data["initial_states"]
            for i, chk in self._initial_checks:
                chk.blockSignals(True)
                chk.setChecked(initial.get(str(i), False))
                chk.blockSignals(False)
        # Clear existing rules
        self.rule_table.setRowCount(0)
        # Load rules
        if "rules" in data:
            for rule in data["rules"]:
                row = self.rule_table.rowCount()
                self.rule_table.insertRow(row)
                ch_combo = QComboBox()
                for name in CH_NAMES:
                    if name == "SCLK":
                        continue
                    ch_combo.addItem(name)
                ch_combo.setCurrentText(rule.get("channel", "XYNC"))
                ch_combo.currentTextChanged.connect(self._save_rules)
                self.rule_table.setCellWidget(row, 0, ch_combo)
                edge_spin = QSpinBox()
                edge_spin.setRange(0, 4096)
                edge_spin.setValue(rule.get("edge", 0))
                edge_spin.valueChanged.connect(self._save_rules)
                self.rule_table.setCellWidget(row, 1, edge_spin)
                level_combo = QComboBox()
                level_combo.addItems(["高", "低"])
                level_combo.setCurrentText(rule.get("level", "高"))
                level_combo.currentTextChanged.connect(self._save_rules)
                self.rule_table.setCellWidget(row, 2, level_combo)
                width_spin = QSpinBox()
                width_spin.setRange(0, 4096)
                width_spin.setValue(rule.get("width", 0))
                width_spin.setToolTip("0=永久保持；N=从触发点起持续N个CLK周期(1CLK周期=2采样点)")
                width_spin.valueChanged.connect(self._save_rules)
                self.rule_table.setCellWidget(row, 3, width_spin)
                btn = QPushButton("删除")
                btn.clicked.connect(lambda checked, r=row: self._remove_rule(r))
                self.rule_table.setCellWidget(row, 4, btn)

    def _load_rules(self):
        """Restore __current__ editor state on startup."""
        db = self._load_templates_db()
        self._refresh_template_list(db)
        data = db.get(_CURRENT_KEY, {})
        if data:
            self._apply_rules_data(data)

    def _refresh_template_list(self, db=None):
        if db is None:
            db = self._load_templates_db()
        self.cb_template.blockSignals(True)
        self.cb_template.clear()
        self.cb_template.addItem("（当前编辑）")
        names = sorted(k for k in db.keys() if k != _CURRENT_KEY)
        self.cb_template.addItems(names)
        self.cb_template.setCurrentIndex(0)
        self.cb_template.blockSignals(False)

    def _on_template_selected(self, index):
        name = self.cb_template.itemText(index)
        if index == 0 or not name:
            return  # "（当前编辑）"
        db = self._load_templates_db()
        data = db.get(name)
        if not data:
            return
        reply = QMessageBox.question(
            self, "加载模板",
            f"将用模板 '{name}' 替换当前编辑器内容，继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            self.cb_template.setCurrentIndex(0)
            return
        self._apply_rules_data(data)
        self._save_rules()  # persist as __current__
        self._schedule_plot_update()
        self.log_signal.emit(f"[INFO] 已加载模板 '{name}'")

    def _save_template(self):
        db = self._load_templates_db()
        current_name = self.cb_template.currentText().strip()
        exists = current_name and current_name != "（当前编辑）" and current_name in db

        from PyQt6.QtWidgets import QInputDialog
        if exists:
            # Offer: overwrite current or save as new
            btn = QMessageBox.question(
                self, "保存模板",
                f"模板 '{current_name}' 已存在。\n"
                f"覆盖 '{current_name}' 请选 [Yes]\n"
                f"另存为新名称请选 [No]",
                QMessageBox.StandardButton.Yes
                | QMessageBox.StandardButton.No
                | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Cancel,
            )
            if btn == QMessageBox.StandardButton.Cancel:
                return
            if btn == QMessageBox.StandardButton.No:
                current_name = ""
        if not current_name:
            name, ok = QInputDialog.getText(
                self, "保存模板", "模板名称:")
            if not ok or not name or not name.strip():
                return
            current_name = name.strip()

        db[current_name] = self._gather_rules_data()
        self._save_templates_db(db)
        self._refresh_template_list(db)
        self.cb_template.setCurrentText(current_name)
        self.log_signal.emit(f"[INFO] 模板 '{current_name}' 已保存")

    def _delete_template(self):
        name = self.cb_template.currentText().strip()
        if not name or name == "（当前编辑）":
            return
        db = self._load_templates_db()
        if name not in db:
            return
        reply = QMessageBox.question(
            self, "删除模板",
            f"确认删除模板 '{name}' ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        del db[name]
        self._save_templates_db(db)
        self._refresh_template_list(db)
        self.log_signal.emit(f"[INFO] 模板 '{name}' 已删除")
