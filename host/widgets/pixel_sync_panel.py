# -*- coding: utf-8 -*-
"""Pixel-sync capture panel: PA6 ADC + XYNC/CLK GPIO flags → Excel.

Captures PA6 (VOUTS) using the existing ADC DMA burst path.  At the
same time the firmware samples PA0 (XYNC) and PA1 (CLK) in the TIM3
update ISR, then sends the flags as CMD_PIXEL_DATA.
Results are exported to Excel for offline pixel marking.
"""

import os
import threading
from datetime import datetime

import numpy as np
import openpyxl
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QSpinBox, QGroupBox, QMessageBox, QFileDialog,
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal

from comm.protocol import (
    build_adc_config, build_adc_burst, parse_adc_data, parse_pixel_data,
    CMD_ACK, CMD_ADC_DATA, CMD_PIXEL_DATA,
)
from comm.serial_link import SerialLink

ADC_VREF = 3.3
ADC_MAX  = 4095.0


class PixelSyncPanel(QWidget):
    log_signal = pyqtSignal(str)

    def __init__(self, link: SerialLink, parent=None):
        super().__init__(parent)
        self.link = link

        self._lock = threading.Lock()
        self._capture_busy = False
        self._adc_samples = []   # list of uint16 raw PA6 values
        self._flags = []         # list of flags (bit0=XYNC, bit1=CLK)
        self._pending_expect = 0
        self._pending_adc_count = 0
        self._pending_flag_count = 0
        self._pending_adc_buf = []
        self._pending_flag_buf = []

        self._setup_ui()

    # ----------------------------------------------------------------
    # UI
    # ----------------------------------------------------------------

    def _setup_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(8, 8, 8, 8)

        # ── Config row ─────────────────────────────────────────
        cfg = QHBoxLayout()
        cfg.addWidget(QLabel("采样率(Hz):"))
        self.sb_rate = QSpinBox()
        self.sb_rate.setRange(1000, 1000000)
        self.sb_rate.setValue(50000)
        self.sb_rate.setSingleStep(1000)
        cfg.addWidget(self.sb_rate)

        cfg.addWidget(QLabel("样本数:"))
        self.sb_samples = QSpinBox()
        self.sb_samples.setRange(100, 16384)
        self.sb_samples.setValue(5000)
        self.sb_samples.setSingleStep(1000)
        cfg.addWidget(self.sb_samples)

        cfg.addStretch()

        self.btn_capture = QPushButton("采集 XYNC+CLK+VOUTS")
        self.btn_capture.setStyleSheet(
            "font-weight: bold; min-height: 32px; font-size: 14px;")
        self.btn_capture.clicked.connect(self._on_capture)
        cfg.addWidget(self.btn_capture)

        self.btn_export = QPushButton("导出Excel")
        self.btn_export.clicked.connect(self._on_export)
        self.btn_export.setEnabled(False)
        cfg.addWidget(self.btn_export)

        self.btn_compress = QPushButton("压缩CLK波形")
        self.btn_compress.setToolTip(
            "将原始数据按 CLK 半周期合并成 0101 格式")
        self.btn_compress.clicked.connect(self._on_compress)
        self.btn_compress.setEnabled(False)
        cfg.addWidget(self.btn_compress)

        outer.addLayout(cfg)

        # ── Status row ─────────────────────────────────────────
        status_row = QHBoxLayout()
        self.lbl_status = QLabel(
            "就绪 — 将采集 PA6(VOUTS) + XYNC(PA0) + CLK(PA1) 同步数据")
        self.lbl_status.setStyleSheet("color: gray;")
        status_row.addWidget(self.lbl_status)
        status_row.addStretch()
        outer.addLayout(status_row)

        # ── Info group ─────────────────────────────────────────
        info = QGroupBox("说明")
        info_layout = QVBoxLayout(info)
        info_layout.addWidget(QLabel(
            "• 本页面使用 ADC Burst 采集 PA6 电压\n"
            "• 固件在 TIM3 中断里同步读取 PA0 (XYNC) 和 PA1 (CLK) 电平\n"
            "• 采集完成后导出 Excel，可在 Excel 中筛选 XYNC/CLK 标记像素位置\n"
            "• 采样率通常设为 CLK 频率的 10~50 倍以获得足够的像素内细节"
        ))
        outer.addWidget(info)

        outer.addStretch()

    # ----------------------------------------------------------------
    # Capture flow
    # ----------------------------------------------------------------

    def _on_capture(self):
        if not self.link.is_open():
            QMessageBox.warning(self, "未连接", "请先连接串口。")
            return
        if self._capture_busy:
            return

        rate = self.sb_rate.value()
        ns = self.sb_samples.value()

        self._capture_busy = True
        self._pending_expect = ns
        self._pending_adc_count = 0
        self._pending_flag_count = 0
        self._pending_adc_buf = []
        self._pending_flag_buf = []

        self.btn_capture.setEnabled(False)
        self.btn_export.setEnabled(False)
        self.lbl_status.setText(
            f"采集中... {ns} 样本 @ {rate} Hz — 请等待")
        self.lbl_status.setStyleSheet("color: orange; font-weight: bold;")

        # Standard ADC burst, ch 0 = PA6 only
        self.link.send(build_adc_config(0x01, rate, mode=0).to_bytes())
        QTimer.singleShot(50, lambda: self._send_burst(ns))

        # Timeout watchdog: 15 s
        QTimer.singleShot(15000, self._on_timeout)

    def _send_burst(self, ns):
        self.link.send(build_adc_burst(0x01, ns).to_bytes())
        self.log_signal.emit(
            f"[TX] BURST (pixel) ch=0x01 n={ns} rate={self.sb_rate.value()}Hz")

    def _on_timeout(self):
        if not self._capture_busy:
            return
        self._capture_busy = False
        self.btn_capture.setEnabled(True)
        self.lbl_status.setText("采集超时 — 请检查连接后重试")
        self.lbl_status.setStyleSheet("color: red; font-weight: bold;")
        self.log_signal.emit("[WARN] Pixel采集超时(15s)")

    # ----------------------------------------------------------------
    # Frame handler
    # ----------------------------------------------------------------

    def on_frame(self, frame):
        if frame.cmd == CMD_ADC_DATA:
            try:
                seq_id, ch_mask, _, raw = parse_adc_data(frame.payload)
            except ValueError:
                return
            if ch_mask != 0x01:
                return
            samples = np.frombuffer(raw, dtype=np.uint16)
            self._pending_adc_buf.append(samples)
            self._pending_adc_count += len(samples)
            self._check_done()

        elif frame.cmd == CMD_PIXEL_DATA:
            try:
                seq_id, flags = parse_pixel_data(frame.payload)
            except ValueError:
                return
            self._pending_flag_buf.extend(flags)
            self._pending_flag_count += len(flags)
            self._check_done()

        elif frame.cmd == CMD_ACK:
            pass

    def _check_done(self):
        if (self._pending_adc_count >= self._pending_expect and
                self._pending_flag_count >= self._pending_expect):
            self._on_capture_done()

    def _on_capture_done(self):
        self._capture_busy = False

        with self._lock:
            adc = np.concatenate(self._pending_adc_buf) if self._pending_adc_buf else np.array([], dtype=np.uint16)
            self._adc_samples = list(adc[:self._pending_expect])
            self._flags = self._pending_flag_buf[:self._pending_expect]

        self._pending_adc_buf = []
        self._pending_flag_buf = []

        n = len(self._adc_samples)
        self.btn_capture.setEnabled(True)
        self.btn_export.setEnabled(True)
        self.btn_compress.setEnabled(True)
        self.lbl_status.setText(
            f"完成: {n} 样本 | 可直接导出 Excel 或压缩 CLK 波形")
        self.lbl_status.setStyleSheet("color: green; font-weight: bold;")
        self.log_signal.emit(f"[INFO] Pixel采集完成: {n} 样本")

        # Auto-export
        self._export_excel()

    # ----------------------------------------------------------------
    # Excel export
    # ----------------------------------------------------------------

    def _on_export(self):
        if not self._adc_samples:
            QMessageBox.warning(self, "无数据", "请先采集数据。")
            return
        log_dir = "D:/test/STM32H723ZGT6/host/adc_logs"
        os.makedirs(log_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"pixel_sync_{ts}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出像素同步数据", os.path.join(log_dir, default_name),
            "Excel (*.xlsx)")
        if not path:
            return
        self._save_excel_to(path)

    def _export_excel(self):
        """Auto-export to a fixed path (overwritten each capture)."""
        if not self._adc_samples:
            return
        log_dir = "D:/test/STM32H723ZGT6/host/adc_logs"
        os.makedirs(log_dir, exist_ok=True)
        path = os.path.join(log_dir, "pixel_sync_latest.xlsx")
        self._save_excel_to(path)
        self.log_signal.emit(f"[INFO] Excel已保存: {path}")

    def _save_excel_to(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PixelSync"

        headers = ["Seq", "XYNC", "CLK", "VOUTS (V)"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)

        for r in range(len(self._adc_samples)):
            adc = self._adc_samples[r]
            flags = self._flags[r] if r < len(self._flags) else 0
            xyNC = 1 if (flags & 0x01) else 0
            clk = 1 if (flags & 0x02) else 0
            volts = round(adc * ADC_VREF / ADC_MAX, 4)

            ws.cell(row=r + 2, column=1, value=r)       # Seq
            ws.cell(row=r + 2, column=2, value=xyNC)    # XYNC
            ws.cell(row=r + 2, column=3, value=clk)     # CLK
            ws.cell(row=r + 2, column=4, value=volts)   # VOUTS

        wb.save(path)

    # ----------------------------------------------------------------
    # CLK half-cycle compression
    # ----------------------------------------------------------------

    def _on_compress(self):
        if not self._adc_samples:
            QMessageBox.warning(self, "无数据", "请先采集数据。")
            return
        log_dir = "D:/test/STM32H723ZGT6/host/adc_logs"
        os.makedirs(log_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"clk_compressed_{ts}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "压缩 CLK 波形", os.path.join(log_dir, default_name),
            "Excel (*.xlsx)")
        if not path:
            return
        self._save_compressed_to(path)

    def _save_compressed_to(self, path):
        """Compress raw samples by CLK half-cycle.

        Each row in the output represents one contiguous CLK half-cycle
        (0 or 1).  XYNC is taken as the majority value within the group,
        VOUTS is averaged.  No pixel numbering is performed.
        """
        if len(self._adc_samples) != len(self._flags):
            self.log_signal.emit("[WARN] ADC 样本数与 flags 不匹配")
            return

        n = len(self._adc_samples)
        xyNC = [1 if (f & 0x01) else 0 for f in self._flags]
        clk  = [1 if (f & 0x02) else 0 for f in self._flags]

        if n == 0:
            self.log_signal.emit("[WARN] 无数据可压缩")
            return

        groups = []  # (start_seq, xync_val, clk_state, avg_adc)
        current_clk = clk[0]
        sum_adc = self._adc_samples[0]
        xync_votes = [0, 0]
        xync_votes[xyNC[0]] += 1
        start_seq = 0
        count = 1

        for i in range(1, n):
            if clk[i] != current_clk:
                # Close current group
                avg = int(round(sum_adc / count)) if count else 0
                groups.append((start_seq, 1 if xync_votes[1] >= xync_votes[0] else 0,
                               current_clk, avg))
                # Start new group
                current_clk = clk[i]
                sum_adc = 0
                xync_votes = [0, 0]
                start_seq = i
                count = 0

            sum_adc += self._adc_samples[i]
            xync_votes[xyNC[i]] += 1
            count += 1

        # Flush final group
        if count > 0:
            avg = int(round(sum_adc / count))
            groups.append((start_seq, 1 if xync_votes[1] >= xync_votes[0] else 0,
                           current_clk, avg))

        if not groups:
            self.log_signal.emit("[WARN] 未提取到有效 CLK 半周期")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CLKCompressed"

        headers = ["Seq", "XYNC", "CLK", "ADC(raw)", "ADC(V)"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)

        for r, (seq, xv, cv, avg_adc) in enumerate(groups):
            volts = round(avg_adc * ADC_VREF / ADC_MAX, 4)
            ws.cell(row=r + 2, column=1, value=seq)
            ws.cell(row=r + 2, column=2, value=xv)
            ws.cell(row=r + 2, column=3, value=cv)
            ws.cell(row=r + 2, column=4, value=avg_adc)
            ws.cell(row=r + 2, column=5, value=volts)

        wb.save(path)
        self.log_signal.emit(
            f"[INFO] 压缩后 CLK 波形已保存: {path} (共 {len(groups)} 个半周期)")
