# -*- coding: utf-8 -*-
"""Snapshot panel — PA6 + PA7 + CLK(PB0) + XYNC(PC4) capture with waveform + Excel.

Independent from the ADC display panel.  Captures PA6 & PA7 (analog) plus
CLK/XYNC (digital flags) in one burst, shows all four waveforms,
and exports a 5-sheet Excel (raw / CLK-compressed / pixel-map / pixels / 8×8).
"""

import json
import os
import shutil
import subprocess
import sys
import threading
from datetime import datetime

import numpy as np
import pyqtgraph as pg
import openpyxl
from PIL import Image
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QPointF
from PyQt6.QtGui import QFont, QColor
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QSpinBox, QDoubleSpinBox, QMessageBox, QCheckBox,
    QSplitter, QFileDialog, QLineEdit, QSizePolicy,
    QStackedWidget, QTableWidget, QTableWidgetItem, QHeaderView,
    QRadioButton, QButtonGroup,
)

from comm.protocol import build_adc_config, build_adc_burst, parse_adc_data
from comm.protocol import CMD_ACK, CMD_ADC_DATA
from comm.serial_link import SerialLink

ADC_VREF   = 3.3
ADC_MAX    = 4095.0
DIFF_MAX   = 2048.0
THR        = ADC_MAX / 2.0   # digital threshold

MODE_SINGLE = 0
MODE_DIFF   = 1

# ── Single-ended: PA6, PA7, CLK, XYNC ──
SE_PINS    = ["PA6", "PA7", "CLK", "XYNC"]
SE_NAMES   = ["PA6 模拟输入", "PA7 模拟输入", "CLK 像素时钟", "XYNC 帧同步"]
SE_PHYS    = [0, 1, 3, 4]
SE_COLORS  = [
    (255, 80, 80),     # PA6  — red
    (80, 200, 80),     # PA7  — green
    (128, 0, 128),     # CLK  — purple
    (255, 165, 0),     # XYNC — orange
]

# ── Differential: Vout_diff(PA6=INP3, PA7=INN3), CLK, XYNC ──
DF_PINS    = ["Vout_diff", "CLK", "XYNC"]
DF_NAMES   = ["Vout=PA6−PA7 差分", "CLK 像素时钟", "XYNC 帧同步"]
DF_PHYS    = [0, 3, 4]
DF_COLORS  = [
    (255, 80, 80),     # Vout_diff — red
    (128, 0, 128),     # CLK  — purple
    (255, 165, 0),     # XYNC — orange
]


def _raw2v(raw, is_diff):
    r = np.asarray(raw, dtype=np.float32)
    if is_diff:
        return (r - 2048.0) * (ADC_VREF / DIFF_MAX)
    return r * (ADC_VREF / ADC_MAX)


def _resolve_config_dir():
    """配置文件目录，兼容源码运行 / PyInstaller 打包两种场景。"""
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)        # host/dist
        host_dir = os.path.dirname(exe_dir)              # host
    else:
        host_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    cfg = os.path.join(host_dir, "config")
    os.makedirs(cfg, exist_ok=True)
    return cfg


_SNAP_CONFIG_DIR = _resolve_config_dir()
_SNAP_CONFIG_PATH = os.path.join(_SNAP_CONFIG_DIR, "snap_config.json")

# 首次运行：把旧位置 (~/.h723_snap_config.json) 的文件迁移过来
_OLD_SNAP = os.path.join(os.path.expanduser("~"), ".h723_snap_config.json")
if os.path.exists(_OLD_SNAP) and not os.path.exists(_SNAP_CONFIG_PATH):
    try:
        shutil.copy2(_OLD_SNAP, _SNAP_CONFIG_PATH)
    except Exception:
        pass

# 8×8 pixel rearrangement map — CIS zigzag readout order
# PIXEL_MAP[row][col] = 0-based pixel index (0..63)
PIXEL_MAP = [
    [63, 62, 59, 58, 55, 54, 51, 50],
    [61, 60, 57, 56, 53, 52, 49, 48],
    [18, 19, 22, 23, 24, 25, 46, 47],
    [16, 17, 20, 21, 26, 27, 44, 45],
    [15, 14, 11, 10, 28, 29, 43, 42],
    [13, 12,  9,  8, 30, 31, 41, 40],
    [ 2,  3,  6,  7, 32, 33, 38, 39],
    [ 0,  1,  4,  5, 34, 35, 36, 37],
]


class SnapViewBox(pg.ViewBox):
    """Pan horizontally; Ctrl+wheel = horiz zoom, wheel = vert zoom."""
    def __init__(self, *args, **kw):
        super().__init__(*args, **kw)
        self.setMouseMode(pg.ViewBox.PanMode)

    def wheelEvent(self, ev, axis=None):
        if ev.modifiers() & Qt.KeyboardModifier.ControlModifier:
            super().wheelEvent(ev, axis=0)
        else:
            super().wheelEvent(ev, axis=1)
        ev.accept()


class SnapPanel(QWidget):
    log_signal = pyqtSignal(str)
    snap_done  = pyqtSignal()

    def __init__(self, link: SerialLink, parent=None):
        super().__init__(parent)
        self.link = link
        self._lock = threading.Lock()
        self._mode = MODE_SINGLE

        # ── State ────────────────────────────────────────────
        self._pending   = False
        self._ch_mask   = 0
        self._ch_list   = []
        self._bufs      = []
        self._expect    = 0
        self._received  = 0
        self._done_spc  = 0
        self._name_edits  = []
        self._name_labels = []

        self._cfg = [SE_PINS, SE_NAMES, SE_PHYS, SE_COLORS]   # PINS, NAMES, PHYS, COLORS for SE
        self._cfg_diff = [DF_PINS, DF_NAMES, DF_PHYS, DF_COLORS]

        self._setup_ui()
        self._load_config()

        self._plot_timer = QTimer(self)
        self._plot_timer.timeout.connect(self._update_plot)
        self._plot_timer.start(100)

        self.snap_done.connect(self._on_snap_done)

        self._burst_timeout = QTimer(self)
        self._burst_timeout.setSingleShot(True)
        self._burst_timeout.timeout.connect(self._on_burst_timeout)

        # Last snapshot's 8×8 pixel data + tiff path (for preview / ImageJ)
        self._pix_raw   = {}    # pixel_index (0..63) → raw ADC value
        self._tiff_path = ""    # path to the most recent exported TIFF
        self._auto_scale = True   # auto-stretch min/max on each snap

    # ----------------------------------------------------------------
    # UI
    # ----------------------------------------------------------------

    def _setup_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(4, 4, 4, 4)

        # ── Mode row ──────────────────────────────────────────
        mr = QHBoxLayout(); mr.addWidget(QLabel("ADC模式:"))
        self._bg = QButtonGroup(self)
        self.radio_single = QRadioButton("单端 (PA6/PA7/CLK/XYNC)")
        self.radio_diff = QRadioButton("差分 (Vout=PA6−PA7, CLK, XYNC)")
        self._bg.addButton(self.radio_single, MODE_SINGLE); self._bg.addButton(self.radio_diff, MODE_DIFF)
        self.radio_single.setChecked(True); self._bg.idToggled.connect(self._on_mode_changed)
        mr.addWidget(self.radio_single); mr.addWidget(self.radio_diff); mr.addStretch()
        outer.addLayout(mr)

        # ── Config row ─────────────────────────────────────────
        cfg = QHBoxLayout()
        cfg.addWidget(QLabel("采样率(Hz):"))
        self.sb_rate = QSpinBox()
        self.sb_rate.setRange(1000, 1000000)
        self.sb_rate.setValue(50000)
        self.sb_rate.setSingleStep(1000)
        cfg.addWidget(self.sb_rate)

        cfg.addWidget(QLabel("样本数:"))
        self.sb_samples = QSpinBox()
        self.sb_samples.setRange(100, 8192)
        self.sb_samples.setValue(5000)
        self.sb_samples.setSingleStep(1000)
        cfg.addWidget(self.sb_samples)

        # XYNC & CLK checkboxes (PA6/PA7 always enabled)
        self.chk_clk = QCheckBox("CLK (PB0)")
        self.chk_clk.setChecked(True)
        cfg.addWidget(self.chk_clk)
        self.chk_xync = QCheckBox("XYNC (PC4)")
        self.chk_xync.setChecked(True)
        cfg.addWidget(self.chk_xync)

        cfg.addStretch()

        self.btn_snap = QPushButton("快照")
        self.btn_snap.setStyleSheet(
            "font-weight: bold; min-height: 32px; font-size: 14px; "
            "background-color: #FF8C00; color: white;"
            "QPushButton:disabled { background-color: #CCC; color: #888; }")
        self.btn_snap.clicked.connect(self._on_snap)
        cfg.addWidget(self.btn_snap)

        self.btn_export = QPushButton("另存为Excel")
        self.btn_export.setToolTip("将当前快照数据另存为时间戳文件")
        self.btn_export.clicked.connect(self._export_excel_as)
        cfg.addWidget(self.btn_export)
        outer.addLayout(cfg)

        # ── Action row ────────────────────────────────────────
        action_row = QHBoxLayout()
        self.lbl_status = QLabel("就绪 — PA6(ADC1_IN3) + PA7(ADC1_IN7) + PB0-CLK(ADC1_IN9) + PC4-XYNC(ADC1_IN4)")
        self.lbl_status.setStyleSheet("color: gray;")
        action_row.addWidget(self.lbl_status)
        action_row.addStretch()
        self.btn_toggle_plots = QPushButton("隐藏波形")
        self.btn_toggle_plots.setCheckable(True)
        self.btn_toggle_plots.setToolTip("折叠/展开四个波形显示区域")
        self.btn_toggle_plots.toggled.connect(self._on_toggle_plots)
        action_row.addWidget(self.btn_toggle_plots)
        self.btn_reset = QPushButton("重置")
        self.btn_reset.clicked.connect(self._on_reset)
        action_row.addWidget(self.btn_reset)
        outer.addLayout(action_row)

        # ── Name row ──────────────────────────────────────────
        name_row = QHBoxLayout()
        name_row.addWidget(QLabel("波形名:"))
        self._name_row = name_row
        self._name_dynamic_widgets = []
        outer.addLayout(name_row)

        # ── 3 linked plots ────────────────────────────────────
        self._plots  = []
        self._curves = []

        splitter = QSplitter(Qt.Orientation.Vertical)

        y_labels = SE_NAMES
        y_ranges = [(0, 4), (0, 4), (-0.1, 1.5), (-0.1, 1.5)]

        for i in range(len(SE_PINS)):
            pw = pg.PlotWidget(viewBox=SnapViewBox())
            pw.setLabel("left", y_labels[i])
            color = SE_COLORS[i]
            pw.getAxis("left").setPen(color)
            pw.getAxis("left").setTextPen(color)
            pw.setYRange(*y_ranges[i], padding=0)
            pw.setXRange(0, 5000)
            pw.showGrid(x=True, y=True)

            pen = pg.mkPen(color=color, width=1.5)
            curve = pw.plot(pen=pen, name=SE_NAMES[i])

            txt = pg.TextItem(
                text=SE_NAMES[i], color=color,
                anchor=(0, 0.5), fill=pg.mkColor(0, 0, 0, 160),
            )
            txt.setFont(QFont("", 11))
            txt.setZValue(10)
            pw.addItem(txt)
            self._name_labels.append(txt)

            self._plots.append(pw)
            self._curves.append(curve)
            splitter.addWidget(pw)

        # Link X-axes
        base_vb = self._plots[0].getViewBox()
        for i in range(1, len(SE_PINS)):
            self._plots[i].getViewBox().setXLink(base_vb)

        self._plots[-1].setLabel("bottom", "样本序号")
        for i in range(len(SE_PINS) - 1):
            self._plots[i].getAxis("bottom").setStyle(showValues=False)

        splitter.setSizes([150, 150, 150, 150])
        self._se_splitter = splitter

        # Build diff-mode splitter (3 plots: Vout_diff, CLK, XYNC)
        self._df_splitter = self._build_diff_splitter()

        # Stacked widget: page 0 = SE waveform plots, page 1 = DF waveform plots,
        # page 2 = 8×8 pixel preview.
        self._stack = QStackedWidget()
        self._stack.addWidget(self._se_splitter)
        self._stack.addWidget(self._df_splitter)

        preview = QVBoxLayout()
        preview.setContentsMargins(8, 8, 8, 8)
        preview.setSpacing(4)
        title_lbl = QLabel("8×8 像素重排 (PA7)")
        title_lbl.setFont(QFont("", 12, QFont.Weight.Bold))
        title_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        preview.addWidget(title_lbl)

        # ── Range bar: min/max spin boxes + auto-scale toggle ──────
        range_row = QHBoxLayout()
        range_row.addWidget(QLabel("范围:"))
        self._spin_min = QSpinBox()
        self._spin_min.setRange(0, 4095)
        self._spin_min.setValue(0)
        self._spin_min.setFixedWidth(90)
        self._spin_min.editingFinished.connect(self._refresh_preview)
        range_row.addWidget(self._spin_min)
        range_row.addWidget(QLabel("~"))
        self._spin_max = QSpinBox()
        self._spin_max.setRange(0, 4095)
        self._spin_max.setValue(4095)
        self._spin_max.setFixedWidth(90)
        self._spin_max.editingFinished.connect(self._refresh_preview)
        range_row.addWidget(self._spin_max)
        self._chk_auto = QCheckBox("自动")
        self._chk_auto.setChecked(True)
        self._chk_auto.toggled.connect(self._on_auto_scale_toggled)
        self._chk_auto.setToolTip("勾选时每次快照自动拉伸范围；"
                                  "取消后可手动设定 min/max")
        range_row.addWidget(self._chk_auto)
        range_row.addStretch()
        preview.addLayout(range_row)

        self._preview_table = QTableWidget(8, 8)
        self._preview_table.setHorizontalHeaderLabels(
            [f"C{c}" for c in range(8)])
        self._preview_table.setVerticalHeaderLabels(
            [f"R{r}" for r in range(8)])
        self._preview_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self._preview_table.verticalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self._preview_table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers)
        self._preview_table.setSelectionMode(
            QTableWidget.SelectionMode.NoSelection)
        self._preview_table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self._preview_table.setToolTip(
            "颜色 = PA7 原始值 (黑=0, 白=4095); self._preview_table 显示像素顺序编号")
        preview.addWidget(self._preview_table, stretch=1)

        self._preview_status = QLabel("尚无快照数据")
        self._preview_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._preview_status.setStyleSheet("color: gray;")
        preview.addWidget(self._preview_status)

        preview_wrap = QWidget()
        preview_wrap.setLayout(preview)
        self._stack.addWidget(preview_wrap)

        outer.addWidget(self._stack, stretch=1)

    def _build_diff_splitter(self):
        splitter = QSplitter(Qt.Orientation.Vertical)
        y_labels = DF_NAMES
        y_ranges = [(-3.4, 3.4), (-0.1, 1.5), (-0.1, 1.5)]
        for i, d in enumerate(DF_PINS):
            pw = pg.PlotWidget(viewBox=SnapViewBox())
            pw.setLabel("left", y_labels[i])
            color = DF_COLORS[i]
            pw.getAxis("left").setPen(color); pw.getAxis("left").setTextPen(color)
            pw.setYRange(*y_ranges[i], padding=0)
            pw.setXRange(0, 5000)
            pw.showGrid(x=True, y=True)
            pw.plot(pen=pg.mkPen(color=color, width=1.5), name=y_labels[i])
            txt = pg.TextItem(text=y_labels[i], color=color, anchor=(0, 0.5),
                              fill=pg.mkColor(0, 0, 0, 160))
            txt.setFont(QFont("", 11)); txt.setZValue(10); pw.addItem(txt)
            splitter.addWidget(pw)
        base_vb = splitter.widget(0).getViewBox()
        for i in range(1, len(DF_PINS)):
            splitter.widget(i).getViewBox().setXLink(base_vb)
        splitter.widget(len(DF_PINS) - 1).setLabel("bottom", "样本序号")
        for i in range(len(DF_PINS) - 1):
            splitter.widget(i).getAxis("bottom").setStyle(showValues=False)
        splitter.setSizes([200, 100, 100])
        return splitter

    # ----------------------------------------------------------------
    # Name / config
    # ----------------------------------------------------------------

    def _on_mode_changed(self, mid, checked):
        if not checked or mid == self._mode: return
        self._mode = mid
        self._rebuild_name_row()
        self._stack.setCurrentIndex(0 if mid == MODE_SINGLE else 1)
        if self.link.is_open():
            self.link.send(build_adc_config(self._build_mask(), self.sb_rate.value(), mode=self._mode).to_bytes())
            self.log_signal.emit(f"[INFO] 快照模式 → {'差分' if mid == MODE_DIFF else '单端'}")

    def _rebuild_name_row(self):
        for w in self._name_dynamic_widgets:
            self._name_row.removeWidget(w); w.deleteLater()
        self._name_dynamic_widgets.clear()
        self._name_edits.clear()
        pins = SE_PINS if self._mode == MODE_SINGLE else DF_PINS
        names = SE_NAMES if self._mode == MODE_SINGLE else DF_NAMES
        colors = SE_COLORS if self._mode == MODE_SINGLE else DF_COLORS
        for i, pin in enumerate(pins):
            color = colors[i]
            pin_lbl = QLabel(pin)
            pin_lbl.setStyleSheet(f"color: rgb({color[0]},{color[1]},{color[2]}); font-weight: bold")
            self._name_row.addWidget(pin_lbl); self._name_dynamic_widgets.append(pin_lbl)
            display_name = names[i].replace(pin, "").lstrip(" |")
            edit = QLineEdit(display_name if display_name else names[i])
            edit.setMaximumWidth(120)
            edit.textChanged.connect(lambda text, ii=i: self._on_name_changed(ii, text))
            self._name_row.addWidget(edit); self._name_dynamic_widgets.append(edit); self._name_edits.append(edit)
        self._name_row.addStretch()

    def _on_name_changed(self, idx: int, text: str):
        pins = SE_PINS if self._mode == MODE_SINGLE else DF_PINS
        if 0 <= idx < len(pins):
            pin = pins[idx]
            full = f"{pin} | {text}" if text.strip() else pin
            if self._mode == MODE_SINGLE:
                SE_NAMES[idx] = full
            else:
                DF_NAMES[idx] = full
            # Update y-axis label on the active splitter
            cur_splitter = self._se_splitter if self._mode == MODE_SINGLE else self._df_splitter
            cur_splitter.widget(idx).setLabel("left", full)
            self._save_config()

    def _save_config(self):
        try:
            cfg = {"mode": self._mode,
                   "se_names": SE_NAMES[:], "df_names": DF_NAMES[:]}
            with open(_SNAP_CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _load_config(self):
        try:
            if os.path.exists(_SNAP_CONFIG_PATH):
                with open(_SNAP_CONFIG_PATH, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                saved_mode = cfg.get("mode", MODE_SINGLE)
                se_names = cfg.get("se_names", SE_NAMES[:])
                df_names = cfg.get("df_names", DF_NAMES[:])
                for i, n in enumerate(se_names[:len(SE_NAMES)]): SE_NAMES[i] = n
                for i, n in enumerate(df_names[:len(DF_NAMES)]): DF_NAMES[i] = n
                self._mode = saved_mode
                (self.radio_diff if saved_mode == MODE_DIFF else self.radio_single).setChecked(True)
                self._rebuild_name_row()
                self._stack.setCurrentIndex(0 if saved_mode == MODE_SINGLE else 1)
        except Exception:
            self._rebuild_name_row()

    # ----------------------------------------------------------------
    # Channel mask
    # ----------------------------------------------------------------

    def _build_mask(self):
        if self._mode == MODE_DIFF:
            mask = 0x01        # PA6=INP3 (diff negative=PA7=INN3, not in mask)
        else:
            mask = 0x01 | 0x02  # PA6 + PA7 single-ended
        if self.chk_clk.isChecked():
            mask |= 0x08       # CLK (PB0)
        if self.chk_xync.isChecked():
            mask |= 0x10       # XYNC (PC4)
        return mask

    # ----------------------------------------------------------------
    # Capture
    # ----------------------------------------------------------------

    def _on_snap(self):
        if not self.link.is_open():
            QMessageBox.warning(self, "未连接", "请先连接串口。")
            return
        if self._pending:
            QMessageBox.warning(self, "采集中", "请等待当前采集完成。")
            return

        mask = self._build_mask()
        phys_list = SE_PHYS if self._mode == MODE_SINGLE else DF_PHYS
        ch_list = [i for i in range(len(phys_list)) if mask & (1 << phys_list[i])]
        num_ch = len(ch_list)
        if num_ch < 2:
            QMessageBox.warning(self, "通道不足",
                                "快照至少需要 2 个通道。")
            return

        rate = self.sb_rate.value()
        ns   = self.sb_samples.value()

        self._ch_mask  = mask
        self._ch_list  = ch_list
        self._bufs     = [[] for _ in range(num_ch)]
        self._expect   = ns * num_ch
        self._received = 0
        self._done_spc = 0

        self.btn_snap.setEnabled(False)
        pins = ", ".join((SE_PINS if self._mode == MODE_SINGLE else DF_PINS)[i] for i in ch_list)
        self.lbl_status.setText(f"快照中... {ns}样本 x {num_ch}通道({pins}) @ {rate}Hz")
        self.lbl_status.setStyleSheet("color: orange; font-weight: bold;")

        self.link.send(build_adc_config(mask, rate, mode=self._mode).to_bytes())
        QTimer.singleShot(50, lambda: self._send_burst(mask, ns))

    def _send_burst(self, mask, ns):
        self.link.send(build_adc_burst(mask, ns).to_bytes())
        self._pending = True
        self._burst_timeout.start(10000)
        self.log_signal.emit(
            f"[TX] SNAP ch=0x{mask:02X} n={ns} "
            f"rate={self.sb_rate.value()}Hz")

    # ----------------------------------------------------------------
    # Reset
    # ----------------------------------------------------------------

    def _on_toggle_plots(self, checked: bool):
        if checked:
            # Show 8×8 pixel preview (page 1), hide waveforms (page 0)
            self._refresh_preview()
            self._stack.setCurrentIndex(1)
            self.btn_toggle_plots.setText("显示波形")
        else:
            self._stack.setCurrentIndex(0)
            self.btn_toggle_plots.setText("隐藏波形")

    def _refresh_preview(self):
        """Redraw the 8×8 grid using the last snapshot's PA7 raw values."""
        tbl = self._preview_table
        pix = self._pix_raw
        if not pix:
            self._preview_status.setText("尚无快照数据")
            for r in range(8):
                for c in range(8):
                    item = QTableWidgetItem("")
                    item.setBackground(QColor(240, 240, 240))
                    tbl.setItem(r, c, item)
            return
        vals = list(pix.values())
        if self._auto_scale:
            lo, hi = min(vals), max(vals)
            # reflect into the spin boxes (without re-triggering refresh)
            self._spin_min.blockSignals(True)
            self._spin_max.blockSignals(True)
            self._spin_min.setValue(lo)
            self._spin_max.setValue(hi)
            self._spin_min.blockSignals(False)
            self._spin_max.blockSignals(False)
        else:
            lo = self._spin_min.value()
            hi = self._spin_max.value()
            if lo > hi:
                lo, hi = hi, lo
        rng = max(hi - lo, 1)   # avoid div-by-zero
        self._preview_status.setText(
            f"共 {len(pix)} / 64 像素  |  范围 {lo} ~ {hi}  "
            f"|  TIFF: …{self._tiff_path[-40:]}")
        for r in range(8):
            for c in range(8):
                idx = PIXEL_MAP[r][c]
                raw = pix.get(idx, 0)
                item = QTableWidgetItem(str(raw))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                # scale: lo → black, hi → white
                g = max(0, min(255, int((raw - lo) * 255 / rng)))
                item.setBackground(QColor(g, g, g))
                # contrasting text
                item.setForeground(
                    QColor(255, 255, 255) if g < 128 else QColor(0, 0, 0))
                tbl.setItem(r, c, item)

    def _on_auto_scale_toggled(self, checked: bool):
        self._auto_scale = checked
        self._spin_min.setEnabled(not checked)
        self._spin_max.setEnabled(not checked)
        # Spin boxes disabled in auto mode (read-out only)
        self._refresh_preview()

    def _on_open_imagej(self):
        path = getattr(self, "_tiff_path", "")
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "无 TIFF",
                                "请先执行快照采集。")
            return
        exe = self._find_imagej()
        if exe:
            subprocess.Popen([exe, path])
            self.log_signal.emit(f"[INFO] 用 ImageJ 打开: {path}")
        else:
            # Fall back to system default handler for .tif
            os.startfile(path)
            self.log_signal.emit(
                "[INFO] ImageJ 未找到，用系统默认程序打开 TIFF")

    def _find_imagej(self):
        """Return the path to an ImageJ / Fiji executable, or '' if none."""
        # 1) Explicit env override
        env = os.environ.get("IMAGEJ_EXE", "")
        if env and os.path.isfile(env):
            return env
        # 2) Search PATH via shutil.which
        for name in ("ImageJ.exe", "ImageJ-win64.exe", "fiji-linux64"):
            found = shutil.which(name)
            if found:
                return found
        # 3) Common install locations
        candidates = [
            r"C:\Program Files\ImageJ\ImageJ.exe",
            r"C:\Program Files (x86)\ImageJ\ImageJ.exe",
            r"C:\Program Files\Fiji.app\ImageJ-win64.exe",
            r"C:\Program Files (x86)\Fiji.app\ImageJ-win64.exe",
            r"C:\Fiji.app\ImageJ-win64.exe",
        ]
        for p in candidates:
            if os.path.isfile(p):
                return p
        return ""

    def _on_reset(self):
        self._pending = False
        self._burst_timeout.stop()
        self._bufs = []
        self.btn_snap.setEnabled(True)
        for sp in (self._se_splitter, self._df_splitter):
            for i in range(sp.count()):
                for item in sp.widget(i).listDataItems():
                    item.setData([], [])
        self.lbl_status.setText("已重置")
        self.lbl_status.setStyleSheet("color: gray;")
        self.log_signal.emit("[INFO] 快照已重置")

    # ----------------------------------------------------------------
    # Completion
    # ----------------------------------------------------------------

    def _on_snap_done(self):
        self._update_plot()
        self._export_snap_excel()
        # Always refresh the 8×8 preview so it updates even if the
        # user is currently viewing page 1 (8×8) instead of page 0 (plots)
        self._refresh_preview()
        spc = self._done_spc
        self.btn_snap.setEnabled(True)
        self.lbl_status.setText(f"快照完成: {spc} 样本")
        self.lbl_status.setStyleSheet("color: green; font-weight: bold;")
        self.log_signal.emit(f"[INFO] 快照完成: {spc} spc")

    def _on_burst_timeout(self):
        if not self._pending:
            return
        self._pending = False
        self.btn_snap.setEnabled(True)
        self.lbl_status.setText("采集超时，数据可能丢失")
        self.lbl_status.setStyleSheet("color: red; font-weight: bold;")
        self.log_signal.emit("[WARN] 快照超时(10s)，已自动重置")

    # ----------------------------------------------------------------
    # Frame handler (called from serial reader thread)
    # ----------------------------------------------------------------

    def on_frame(self, frame):
        if frame.cmd != CMD_ADC_DATA:
            return
        try:
            seq_id, ch_mask, _, raw = parse_adc_data(frame.payload)
        except ValueError:
            return
        samples = np.frombuffer(raw, dtype=np.uint16)
        if len(samples) == 0:
            return
        num_en = bin(ch_mask).count("1")
        if num_en == 0:
            return
        scan_periods = len(samples) // num_en
        if scan_periods == 0:
            return
        samples = samples[:scan_periods * num_en]
        reshaped = samples.reshape(scan_periods, num_en)

        if self._pending and ch_mask == self._ch_mask:
            for c in range(num_en):
                self._bufs[c].append(reshaped[:, c].copy())
            self._received += scan_periods
            expected_spc = self._expect // num_en

            if self._received >= expected_spc:
                for c in range(num_en):
                    self._bufs[c] = (
                        list(np.concatenate(self._bufs[c]))
                        if self._bufs[c] else [])
                self._pending = False
                self._burst_timeout.stop()
                self._done_spc = expected_spc
                self.snap_done.emit()

    # ----------------------------------------------------------------
    # Plot update
    # ----------------------------------------------------------------

    def _update_plot(self):
        bufs = self._bufs
        ch_list = self._ch_list
        if not ch_list or not bufs:
            return
        phys_list = SE_PHYS if self._mode == MODE_SINGLE else DF_PHYS
        splitter = self._se_splitter if self._mode == MODE_SINGLE else self._df_splitter

        for col, plot_idx in enumerate(ch_list):
            # find the curve in this splitter page
            curves = splitter.widget(plot_idx).listDataItems()
            if not curves:
                continue
            curve = curves[0]
            buf = bufs[col] if col < len(bufs) else []
            if not buf:
                curve.setData([], [])
                continue
            phys = phys_list[plot_idx]
            n = len(buf)

            if self._mode == MODE_SINGLE:
                # SE: PA6/PA7 analog (phys 0,1); CLK/XYNC digital (phys 3,4)
                if phys in (0, 1):
                    volts = _raw2v(buf, False)
                else:
                    volts = np.where(np.asarray(buf) > THR, 1.0, 0.0).astype(np.float32)
            else:
                # DF: Vout_diff analog-diff (phys 0); CLK/XYNC digital (phys 3,4)
                if phys == 0:
                    volts = _raw2v(buf, True)
                else:
                    volts = np.where(np.asarray(buf) > THR, 1.0, 0.0).astype(np.float32)

            if n > 16384:
                bs = max(n // 4000, 2)
                trim = n - (n % bs)
                v = volts[:trim].reshape(-1, bs)
                v_max = v.max(axis=1); v_min = v.min(axis=1)
                x = np.arange(0, trim, bs, dtype=np.float64)
                x2 = np.empty(len(v_max) * 2, dtype=np.float64)
                y2 = np.empty(len(v_max) * 2, dtype=np.float32)
                x2[0::2] = x;  x2[1::2] = x + bs
                y2[0::2] = v_min; y2[1::2] = v_max
                curve.setData(x2, y2)
            else:
                curve.setData(np.arange(n, dtype=np.float64), volts)

    # ----------------------------------------------------------------
    # Excel export
    # ----------------------------------------------------------------

    def _export_snap_excel(self):
        ch_list = self._ch_list
        if not ch_list:
            self.log_signal.emit("[WARN] 没有快照数据")
            return
        log_dir = r"D:\test\STM32H723ZGT6\host\adc_logs"
        os.makedirs(log_dir, exist_ok=True)
        # 固定文件名，每次快照自动覆盖（另存为Excel不受影响）
        path = os.path.join(log_dir, "snap_latest.xlsx")
        self._save_excel_to(path)

    def _export_excel_as(self):
        if not self._bufs or not self._ch_list:
            QMessageBox.warning(self, "无数据", "请先执行快照采集。")
            return
        log_dir = r"D:\test\STM32H723ZGT6\host\adc_logs"
        os.makedirs(log_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"snap_{ts}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "另存为", os.path.join(log_dir, default_name),
            "Excel (*.xlsx)")
        if not path:
            return
        self._save_excel_to(path)

    def _save_excel_to(self, path):
        ch_list = self._ch_list
        bufs = self._bufs
        num_ch = len(ch_list)
        spc = min(len(bufs[c]) for c in range(num_ch)) if num_ch else 0
        if spc == 0:
            return

        phys_list = SE_PHYS if self._mode == MODE_SINGLE else DF_PHYS
        pins = SE_PINS if self._mode == MODE_SINGLE else DF_PINS
        col_phys = [phys_list[i] for i in ch_list]
        col_pa6  = next((c for c, p in enumerate(col_phys) if p == 0), None)
        col_pa7  = next((c for c, p in enumerate(col_phys) if p == 1), None)
        col_clk  = next((c for c, p in enumerate(col_phys) if p == 3), None)
        col_xync = next((c for c, p in enumerate(col_phys) if p == 4), None)

        wb = openpyxl.Workbook()

        # ── Sheet 1: raw data ──────────────────────────────────
        ws = wb.active
        ws.title = "原始数据"
        headers = ["样本序号"]
        for phys in col_phys:
            if phys == 0 and self._mode == MODE_DIFF:
                headers.append("Vout_diff 差分(V)")
            elif phys == 0:
                headers.append("PA6 (V)")
            elif phys == 1:
                headers.append("PA7 (V)")
            elif phys == 3:
                headers.append("CLK")
            elif phys == 4:
                headers.append("XYNC")
            else:
                headers.append(f"CH{phys} (raw)")
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)

        for r in range(spc):
            ws.cell(row=r + 2, column=1, value=r)
            for c in range(num_ch):
                phys = col_phys[c]
                raw = bufs[c][r]
                if phys == 0 and self._mode == MODE_DIFF:
                    ws.cell(row=r + 2, column=c + 2,
                            value=round(_raw2v(raw, True), 4))
                elif phys in (0, 1):
                    ws.cell(row=r + 2, column=c + 2,
                            value=round(raw * ADC_VREF / ADC_MAX, 4))
                elif phys in (3, 4):
                    ws.cell(row=r + 2, column=c + 2,
                            value=1 if raw > THR else 0)
                else:
                    ws.cell(row=r + 2, column=c + 2, value=int(raw))

        # ── Sheet 2–4: CLK compression & pixel extraction ──────
        if col_clk is not None:
            # Build digital arrays
            clk_arr  = [1 if bufs[col_clk][r] > THR else 0 for r in range(spc)]
            xync_arr = [1 if bufs[col_xync][r] > THR else 0 for r in range(spc)] if col_xync is not None else None
            pa6_arr  = [bufs[col_pa6][r] for r in range(spc)] if col_pa6 is not None else None
            pa7_arr  = [bufs[col_pa7][r] for r in range(spc)] if col_pa7 is not None else None
            # DF mode: use Vout_diff (phys=0) as the analog channel for pixel extraction
            diff_arr = [bufs[col_pa6][r] for r in range(spc)] if (self._mode == MODE_DIFF and col_pa6 is not None) else None

            # groups: (clk, xync, pa6_avg_raw, pa7_avg_raw, start_sample, sample_count)
            groups = []
            cur_clk = clk_arr[0]
            start_s = 0
            xync_sum = xync_arr[0] if xync_arr else 0
            pa6_sum  = pa6_arr[0] if pa6_arr else 0
            pa7_sum  = pa7_arr[0] if pa7_arr else 0
            cnt = 1

            for r in range(1, spc):
                if clk_arr[r] != cur_clk:
                    xync_val = 1 if xync_arr and xync_sum * 2 >= cnt else 0
                    pa6_avg  = int(round(pa6_sum / cnt)) if pa6_arr else 0
                    pa7_avg  = int(round(pa7_sum / cnt)) if pa7_arr else 0
                    groups.append((cur_clk, xync_val, pa6_avg, pa7_avg, start_s, cnt))
                    cur_clk = clk_arr[r]; start_s = r
                    xync_sum = xync_arr[r] if xync_arr else 0
                    pa6_sum  = pa6_arr[r] if pa6_arr else 0
                    pa7_sum  = pa7_arr[r] if pa7_arr else 0
                    cnt = 1
                else:
                    if xync_arr: xync_sum += xync_arr[r]
                    if pa6_arr:  pa6_sum  += pa6_arr[r]
                    if pa7_arr:  pa7_sum  += pa7_arr[r]
                    cnt += 1
            if cnt > 0:
                xync_val = 1 if xync_arr and xync_sum * 2 >= cnt else 0
                pa6_avg  = int(round(pa6_sum / cnt)) if pa6_arr else 0
                pa7_avg  = int(round(pa7_sum / cnt)) if pa7_arr else 0
                groups.append((cur_clk, xync_val, pa6_avg, pa7_avg, start_s, cnt))

            # Sheet 2: CLK压缩
            ws2 = wb.create_sheet("CLK压缩")
            for ci, h in enumerate(
                ["半周期序号", "起始样本", "样本数", "CLK", "XYNC",
                 "PA6均值(V)", "PA7均值(V)"], 1):
                ws2.cell(row=1, column=ci, value=h)
            for gi, (clk_s, xync_v, pa6_a, pa7_a, ss, cnt) in enumerate(groups):
                row = gi + 2
                ws2.cell(row=row, column=1, value=gi)
                ws2.cell(row=row, column=2, value=ss)
                ws2.cell(row=row, column=3, value=cnt)
                ws2.cell(row=row, column=4, value=clk_s)
                ws2.cell(row=row, column=5, value=xync_v)
                ws2.cell(row=row, column=6,
                         value=round(pa6_a * ADC_VREF / ADC_MAX, 4))
                ws2.cell(row=row, column=7,
                         value=round(pa7_a * ADC_VREF / ADC_MAX, 4))

            self.log_signal.emit(
                f"[INFO] CLK压缩: {spc} 采样 → {len(groups)} 半周期")

            # Pixel extraction
            if col_xync is not None and col_pa6 is not None:
                xync1_start = next((i for i, g in enumerate(groups) if g[1] == 1), None)
                if xync1_start is not None:
                    pixel1_idx = None
                    i = xync1_start
                    while i < len(groups) and groups[i][1] == 1:
                        if groups[i][0] == 1:
                            pixel1_idx = i
                        i += 1
                    xync1_end = i

                    if pixel1_idx is not None:
                        pixels = []
                        pixels.append((1, pixel1_idx) + groups[pixel1_idx])
                        pixel_num = 2
                        for j in range(xync1_end, len(groups)):
                            if pixel_num > 64: break
                            if groups[j][0] == 1:
                                pixels.append((pixel_num, j) + groups[j])
                                pixel_num += 1

                        if pixels:
                            # Sheet 3: 像素映射
                            ws3 = wb.create_sheet("像素映射")
                            for ci, h in enumerate(
                                ["Pixel", "半周期序号", "起始样本", "样本数",
                                 "CLK", "XYNC", "PA6 (V)", "PA7 (V)"], 1):
                                ws3.cell(row=1, column=ci, value=h)
                            for pi, (pn, gi, clk_s, xync_v, pa6_a, pa7_a, ss, cnt) in enumerate(pixels):
                                row = pi + 2
                                ws3.cell(row=row, column=1, value=pn)
                                ws3.cell(row=row, column=2, value=gi)
                                ws3.cell(row=row, column=3, value=ss)
                                ws3.cell(row=row, column=4, value=cnt)
                                ws3.cell(row=row, column=5, value=clk_s)
                                ws3.cell(row=row, column=6, value=xync_v)
                                ws3.cell(row=row, column=7,
                                         value=round(pa6_a * ADC_VREF / ADC_MAX, 4))
                                ws3.cell(row=row, column=8,
                                         value=round(pa7_a * ADC_VREF / ADC_MAX, 4))

                            # Sheet 4: Pixel数据
                            ws4 = wb.create_sheet("Pixel数据")
                            ws4.cell(row=1, column=1, value="Pixel")
                            ws4.cell(row=1, column=2, value="PA6 (V)")
                            ws4.cell(row=1, column=3, value="PA6 (raw)")
                            ws4.cell(row=1, column=4, value="PA7 (V)")
                            ws4.cell(row=1, column=5, value="PA7 (raw)")
                            for pi, (pn, gi, clk_s, xync_v, pa6_a, pa7_a, ss, cnt) in enumerate(pixels):
                                ws4.cell(row=pi + 2, column=1, value=pn)
                                ws4.cell(row=pi + 2, column=2,
                                         value=round(pa6_a * ADC_VREF / ADC_MAX, 4))
                                ws4.cell(row=pi + 2, column=3,
                                         value=int(pa6_a))
                                ws4.cell(row=pi + 2, column=4,
                                         value=round(pa7_a * ADC_VREF / ADC_MAX, 4))
                                ws4.cell(row=pi + 2, column=5,
                                         value=int(pa7_a))
                            self.log_signal.emit(
                                f"[INFO] Pixel提取: {len(pixels)} pixels")

                            # Sheet 5: 像素重排8×8 (PA7)
                            pix_raw = {}
                            for pn, gi, clk_s, xync_v, pa6_a, pa7_a, ss, cnt in pixels:
                                pix_raw[pn - 1] = int(pa7_a)

                            ws5 = wb.create_sheet("像素重排8×8")
                            # Column headers
                            for c in range(8):
                                ws5.cell(row=1, column=c + 2,
                                         value=f"Col{c}")
                            for r in range(8):
                                ws5.cell(row=r + 2, column=1, value=r)
                                for c in range(8):
                                    idx = PIXEL_MAP[r][c]
                                    raw = pix_raw.get(idx, 0)
                                    ws5.cell(row=r + 2, column=c + 2,
                                             value=raw)

                            # Generate 16-bit TIFF image (8×8, mode I;16, 0-4095)
                            img = Image.new("I;16", (8, 8))
                            pixels_img = img.load()
                            for r in range(8):
                                for c in range(8):
                                    idx = PIXEL_MAP[r][c]
                                    pixels_img[c, r] = pix_raw.get(idx, 0)
                            tiff_path = path.replace(".xlsx", ".tiff")
                            img.save(tiff_path)

                            # Keep references for the preview grid & ImageJ button
                            self._pix_raw = dict(pix_raw)
                            self._tiff_path = tiff_path
                            self.log_signal.emit(
                                f"[INFO] 16-bit TIFF 已保存: {tiff_path}")

        wb.save(path)
        self.log_signal.emit(f"[INFO] 快照 Excel 已保存: {path}")
