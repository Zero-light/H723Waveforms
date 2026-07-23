# -*- coding: utf-8 -*-
"""ADC burst capture — single-ended (PA6/PA7/PC1) or differential (Vout=PA6−PA7, RST=PC1).
Runtime mode switch via radio buttons.
"""

import json, os, shutil, sys, threading
from datetime import datetime
import numpy as np
import pyqtgraph as pg
import openpyxl
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, pyqtSlot, QPointF
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QSpinBox,
    QDoubleSpinBox, QCheckBox, QSplitter, QFileDialog, QLineEdit,
    QStackedWidget, QRadioButton, QButtonGroup,
)
from comm.protocol import build_adc_config, build_adc_burst, parse_adc_data
from comm.protocol import CMD_ACK, CMD_ADC_DATA
from comm.serial_link import SerialLink

CH_COLORS = [(255, 80, 80), (80, 200, 80), (80, 120, 255)]
ADC_VREF = 3.3
ADC_MAX = 4095.0
DIFF_MAX = 2048.0

MODE_SINGLE = 0
MODE_DIFF = 1

# dual voltage markers: ① cyan solid, ② magenta dashed
MK_COLORS = [(0, 255, 255), (255, 110, 255)]
MK_LABELS = ["①", "②"]
MK_STYLES = [Qt.PenStyle.SolidLine, Qt.PenStyle.DashLine]

SE_CH = [
    {"name": "PA6", "pin": "PA6", "diff": False, "color": CH_COLORS[0], "phys": 0},
    {"name": "PA7", "pin": "PA7", "diff": False, "color": CH_COLORS[1], "phys": 1},
    {"name": "PC1", "pin": "PC1", "diff": False, "color": CH_COLORS[2], "phys": 2},
]
DF_CH = [
    {"name": "Vout_diff", "pin": "PA6-7", "diff": True,  "color": CH_COLORS[0], "phys": 0},
    {"name": "RST",       "pin": "PC1",   "diff": False, "color": CH_COLORS[2], "phys": 2},
]


def _raw2v(raw, is_diff):
    """raw: scalar, list, or ndarray; returns same-shape float volts."""
    r = np.asarray(raw, dtype=np.float32)
    if is_diff:
        return (r - 2048.0) * (ADC_VREF / DIFF_MAX)
    return r * (ADC_VREF / ADC_MAX)


def _cfg_dir():
    if getattr(sys, 'frozen', False):
        host_dir = os.path.dirname(os.path.dirname(sys.executable))
    else:
        host_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    d = os.path.join(host_dir, "config")
    os.makedirs(d, exist_ok=True)
    return d

_CFG_DIR = _cfg_dir()
_CFG_PATH = os.path.join(_CFG_DIR, "adc_config.json")
_OLD = os.path.join(os.path.expanduser("~"), ".h723_adc_config.json")
if os.path.exists(_OLD) and not os.path.exists(_CFG_PATH):
    try: shutil.copy2(_OLD, _CFG_PATH)
    except Exception: pass


class _VB(pg.ViewBox):
    def __init__(self, *a, **k):
        super().__init__(*a, **k); self.setMouseMode(pg.ViewBox.PanMode)
    def wheelEvent(self, ev, axis=None):
        ax = 0 if ev.modifiers() & Qt.KeyboardModifier.ControlModifier else axis
        super().wheelEvent(ev, ax)
        ev.accept()


def _build_independent(defs):
    sp = QSplitter(Qt.Orientation.Vertical)
    plots, curves, lbls = [], [], []
    for d in defs:
        pw = pg.PlotWidget(viewBox=_VB())
        pw.setLabel("left", d["name"], units="V")
        c = d["color"]; pw.getAxis("left").setPen(c); pw.getAxis("left").setTextPen(c)
        pw.setYRange(-3.4, 3.4, padding=0) if d["diff"] else pw.setYRange(0, 4, padding=0)
        pw.setXRange(0, 1000, padding=0); pw.enableAutoRange(axis='x', enable=False)
        pw.showGrid(x=True, y=True)
        curve = pw.plot(pen=pg.mkPen(color=c, width=1.5), name=d["name"])
        txt = pg.TextItem(text=d["name"], color=c, anchor=(0, 0.5), fill=pg.mkColor(0, 0, 0, 160))
        txt.setFont(QFont("", 11)); txt.setZValue(10); pw.addItem(txt)
        plots.append(pw); curves.append(curve); lbls.append(txt); sp.addWidget(pw)
    for pw in plots[1:]: pw.getViewBox().setXLink(plots[0].getViewBox())
    plots[-1].setLabel("bottom", "样本序号")
    for pw in plots[:-1]: pw.getAxis("bottom").setStyle(showValues=False)
    return sp, plots, curves, lbls


class AdcPanel(QWidget):
    log_signal = pyqtSignal(str)
    _burst_done = pyqtSignal()

    def __init__(self, link, parent=None):
        super().__init__(parent)
        self.link = link
        self._sr = 20000; self._offs = 0.0; self._lock = threading.Lock()
        self._spins = []; self._edits = []; self._chks = []
        self._mode = MODE_SINGLE
        self._ch_defs = list(SE_CH)
        self._n = len(self._ch_defs)
        self._buf = [[] for _ in range(self._n)]     # single capture buffer
        self._bs = 0; self._bs_ack = 0; self._bpend = False
        self._bexp = 0; self._brec = 0; self._bnb = 0; self._bmask = 0
        self._bmap = []; self._bbuf = []
        # dual voltage markers (①②) per independent plot: each entry is a length-2 list
        self._cur_vl = []; self._cur_tx = []           # [plot][marker] -> InfiniteLine / TextItem
        self._cur_px = []; self._cur_pos = []          # [plot][marker] -> sample idx / scene pos
        self._cur_next = []                             # [plot] -> 0/1 which marker to place next
        self._m_cur = []; self._m_txt = []; self._m_cur_vl = []; self._m_cur_tx = []
        self._m_px = [None, None]; self._m_next = 0; self._m_clk = [None, None]
        self._setup_ui(); self._load()
        self._ptimer = QTimer(self); self._ptimer.timeout.connect(self._update_plot); self._ptimer.start(100)
        self._btimer = QTimer(self); self._btimer.setSingleShot(True); self._btimer.timeout.connect(self._bust_timeout)
        self._burst_done.connect(self._bust_slot)

    # ── config persistence ─────────────────────────────────────────
    def _load(self):
        try:
            cfg = json.load(open(_CFG_PATH, encoding="utf-8")) if os.path.exists(_CFG_PATH) else {}
        except Exception: cfg = {}
        self._offs = float(cfg.get("adc_offset_v", 0.0))
        for w, v in ((self.sb_offs, self._offs), (self.sb_rate, cfg.get("sample_rate", 20000)),
                     (self.sb_pts, cfg.get("sample_points", 1000))):
            w.blockSignals(True); w.setValue(v); w.blockSignals(False)
        m = cfg.get("mode", MODE_SINGLE)
        if m not in (MODE_SINGLE, MODE_DIFF): m = MODE_SINGLE
        self._apply(m, cfg, send=False)

    def _save(self):
        try:
            json.dump({"adc_offset_v": self._offs, "sample_rate": self.sb_rate.value(),
                       "sample_points": self.sb_pts.value(), "mode": self._mode,
                       "channel_enabled": [c.isChecked() for c in self._chks],
                       "channel_names": [e.text() for e in self._edits],
                       "channel_offsets": [s.value() for s in self._spins]},
                      open(_CFG_PATH, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
            self.log_signal.emit("[INFO] ADC 配置已保存")
        except Exception: pass

    # ── mode switching ────────────────────────────────────────────
    def _mode_chg(self, mid, checked):
        if checked and mid != self._mode:
            self._apply(mid, send=self.link.is_open())

    def _apply(self, mode, cfg=None, send=True):
        self._mode = mode
        defs = SE_CH if mode == MODE_SINGLE else DF_CH
        self._ch_defs = list(defs); self._n = len(defs)
        self._buf = [[] for _ in range(self._n)]
        (self.radio_diff if mode == MODE_DIFF else self.radio_single).setChecked(True)
        self._rebuild_ch_rows()
        if mode == MODE_SINGLE:
            plots, curves, lbls = self._se_p, self._se_c, self._se_l
        else:
            plots, curves, lbls = self._df_p, self._df_c, self._df_l
        if not self.btn_merge.isChecked():
            self._stack.setCurrentIndex(0 if mode == MODE_SINGLE else 1)
        self._ip, self._ic, self._il = plots, curves, lbls
        for c in curves + (self._m_cur if mode == self._mode else []):
            self._prepare_curve(c)
        self._rebuild_cursors(); self._rebuild_merged(mode)
        en = (cfg.get("channel_enabled", [True] * self._n) if cfg else [True] * self._n)[:self._n]
        nm = (cfg.get("channel_names", [d["name"] for d in defs]) if cfg else [d["name"] for d in defs])[:self._n]
        of = (cfg.get("channel_offsets", [0.0] * self._n) if cfg else [0.0] * self._n)[:self._n]
        for lst, arr in ((self._chks, en), (self._edits, nm), (self._spins, of)):
            for i, w in enumerate(lst):
                if i < len(arr):
                    w.blockSignals(True)
                    if isinstance(w, QCheckBox): w.setChecked(bool(arr[i]))
                    elif isinstance(w, QLineEdit): w.setText(str(arr[i]))
                    else: w.setValue(float(arr[i]))
                    w.blockSignals(False)
        self._sync_names()
        # enable native downsample on every curve (in case rebuild missed any)
        for c in (self._se_c + self._df_c + self._m_cur):
            self._prepare_curve(c)
        if send and self.link.is_open():
            self.link.send(build_adc_config(self._mask(), self.sb_rate.value(), mode=self._mode).to_bytes())
            self.log_signal.emit(f"[INFO] ADC模式切换为：{'差分' if mode == MODE_DIFF else '单端'}")

    def _rebuild_ch_rows(self):
        # remove old dynamic widgets in cfg_row (channel checkboxes + name edits)
        for w in self._ch_dynamic_widgets:
            self._cfg_row.removeWidget(w); w.deleteLater()
        self._ch_dynamic_widgets.clear()
        # remove old dynamic widgets in off_row (offset spinboxes)
        for w in self._off_dynamic_widgets:
            self._off_row_L.removeWidget(w); w.deleteLater()
        self._off_dynamic_widgets.clear()
        self._chks.clear(); self._edits.clear(); self._spins.clear()

        # --- cfg_row: channel checkbox + pin label + name edit (one set per channel)
        L = self._cfg_row
        p_c = [4]   # position after static widgets: 采样率label, sb_rate, 样本数label, sb_pts

        def _cfg_ins(w):
            L.insertWidget(p_c[0], w); self._ch_dynamic_widgets.append(w); p_c[0] += 1

        for i, d in enumerate(self._ch_defs):
            c = QCheckBox(d["pin"]); c.setChecked(True); c.stateChanged.connect(lambda: self._save())
            self._chks.append(c); _cfg_ins(c)
        for i, d in enumerate(self._ch_defs):
            col = d["color"]; lbl = QLabel(d["pin"])
            lbl.setStyleSheet(f"color: rgb({col[0]},{col[1]},{col[2]}); font-weight: bold")
            _cfg_ins(lbl)

        # --- off_row: name edit + offset spinbox (insert right after 修正Label/sb_offs, before stretch)
        L2 = self._off_row_L
        p_o = [2]   # after corr label + sb_offs

        def _off_ins(w):
            L2.insertWidget(p_o[0], w); self._off_dynamic_widgets.append(w); p_o[0] += 1

        for i, d in enumerate(self._ch_defs):
            col = d["color"]
            _off_ins(QLabel(f"<font color='rgb({col[0]},{col[1]},{col[2]})'>{d['pin']}</font>"))
            e = QLineEdit(d["name"]); e.setMaximumWidth(90)
            e.textChanged.connect(lambda t, ii=i: (self._name_chg(ii, t), self._save()))
            self._edits.append(e); _off_ins(e)
            s = QDoubleSpinBox(); s.setRange(-3.0, 3.0); s.setSingleStep(0.1); s.setValue(0.0)
            s.setDecimals(1); s.setFixedWidth(90)
            s.valueChanged.connect(lambda v, ii=i: (self._schedule(), self._save()))
            self._spins.append(s); _off_ins(s)

    def _rebuild_cursors(self):
        # remove old marker items (robust to nested lists)
        for i, pw in enumerate(self._ip):
            for v in (self._cur_vl[i] if i < len(self._cur_vl) else []):
                try: pw.removeItem(v)
                except Exception: pass
            for t in (self._cur_tx[i] if i < len(self._cur_tx) else []):
                try: pw.removeItem(t)
                except Exception: pass
        self._cur_vl = []; self._cur_tx = []; self._cur_px = []; self._cur_pos = []; self._cur_next = []
        for i, pw in enumerate(self._ip):
            vls = []; txs = []
            for m in range(2):
                mc = MK_COLORS[m]
                vl = pg.InfiniteLine(angle=90, movable=False,
                                     pen=pg.mkPen(color=mc, width=1.4, style=MK_STYLES[m]))
                vl.setVisible(False); pw.addItem(vl); vls.append(vl)
                tx = pg.TextItem(text="", color=mc, anchor=(0, 0), fill=pg.mkColor(0, 0, 0, 180))
                tx.setFont(QFont("", 10)); tx.setZValue(11); tx.setVisible(False); pw.addItem(tx); txs.append(tx)
            self._cur_vl.append(vls); self._cur_tx.append(txs)
            self._cur_px.append([None, None]); self._cur_pos.append([None, None]); self._cur_next.append(0)
            sc = pw.getViewBox().scene()
            try: sc.sigMouseClicked.disconnect()
            except Exception: pass
            sc.sigMouseClicked.connect(lambda ev, ix=i: self._clk_ind(ev, ix))
            try: pw.getViewBox().sigRangeChanged.disconnect()
            except Exception: pass
            pw.getViewBox().sigRangeChanged.connect(self._ind_range_chg)

    def _rebuild_merged(self, mode):
        # remove old items (robust to nested cursor lists)
        flat = list(self._m_cur) + list(self._m_txt)
        for vls in self._m_cur_vl: flat.extend(vls)
        for txs in self._m_cur_tx: flat.extend(txs)
        for it in flat:
            try: self._mp.removeItem(it)
            except Exception: pass
        self._m_cur.clear(); self._m_txt.clear(); self._m_cur_vl.clear(); self._m_cur_tx.clear()
        self._m_px = [None, None]; self._m_next = 0; self._m_clk = [None, None]
        defs = SE_CH if mode == MODE_SINGLE else DF_CH
        self._mp.setYRange(-3.4, 3.4, padding=0) if mode == MODE_DIFF else self._mp.setYRange(0, 4, padding=0)
        for i, d in enumerate(defs):
            col = d["color"]
            cv = self._mp.plot(pen=pg.mkPen(color=col, width=1.5), name=d["name"])
            tx = pg.TextItem(text=d["name"], color=col, anchor=(0, 0.5), fill=pg.mkColor(0, 0, 0, 160))
            tx.setFont(QFont("", 11)); tx.setZValue(10); self._mp.addItem(tx)
            self._m_cur.append(cv); self._m_txt.append(tx)
            vls = []; txs = []
            for m in range(2):
                mc = MK_COLORS[m]
                vl = pg.InfiniteLine(angle=90, movable=False,
                                     pen=pg.mkPen(color=mc, width=1.4, style=MK_STYLES[m]))
                vl.setVisible(False); self._mp.addItem(vl); vls.append(vl)
                ct = pg.TextItem(text="", color=mc, anchor=(0, 0), fill=pg.mkColor(0, 0, 0, 180))
                ct.setFont(QFont("", 10)); ct.setZValue(11); ct.setVisible(False); self._mp.addItem(ct); txs.append(ct)
            self._m_cur_vl.append(vls); self._m_cur_tx.append(txs)

    # ── one-time UI setup ─────────────────────────────────────────
    def _setup_ui(self):
        o = QVBoxLayout(self); o.setContentsMargins(4, 4, 4, 4)

        # mode row
        mr = QHBoxLayout(); mr.addWidget(QLabel("ADC模式:"))
        self._bg = QButtonGroup(self)
        self.radio_single = QRadioButton("单端 (PA6/PA7/PC1)")
        self.radio_diff = QRadioButton("差分 (Vout=PA6−PA7, RST=PC1)")
        self._bg.addButton(self.radio_single, MODE_SINGLE); self._bg.addButton(self.radio_diff, MODE_DIFF)
        self.radio_single.setChecked(True); self._bg.idToggled.connect(self._mode_chg)
        mr.addWidget(self.radio_single); mr.addWidget(self.radio_diff); mr.addStretch(); o.addLayout(mr)

        # config row: 采样率 / 样本数 + 通道勾选框 + 按钮（静态+动态混排）
        self._cfg_row = QHBoxLayout()
        self._cfg_row.addWidget(QLabel("采样率(Hz):"))
        self.sb_rate = QSpinBox(); self.sb_rate.setRange(1000, 1000000); self.sb_rate.setValue(20000)
        self.sb_rate.setSingleStep(1000); self.sb_rate.valueChanged.connect(lambda: self._save())
        self._cfg_row.addWidget(self.sb_rate)
        self._cfg_row.addWidget(QLabel("样本数:"))
        self.sb_pts = QSpinBox(); self.sb_pts.setRange(100, 16384); self.sb_pts.setValue(1000)
        self.sb_pts.setSingleStep(1000); self.sb_pts.valueChanged.connect(lambda: self._save())
        self._cfg_row.addWidget(self.sb_pts)
        self._ch_dynamic_widgets = []
        self._cfg_row.addStretch()
        self.btn_go = QPushButton("采集并导出Excel")
        self.btn_go.setStyleSheet("font-weight: bold; min-height: 32px; font-size: 14px;")
        self.btn_go.clicked.connect(self._go); self._cfg_row.addWidget(self.btn_go)
        self.btn_save_as = QPushButton("另存为Excel"); self.btn_save_as.clicked.connect(self._export_as)
        self._cfg_row.addWidget(self.btn_save_as)
        o.addLayout(self._cfg_row)

        # ADC修正 row
        cr = QHBoxLayout(); cr.addWidget(QLabel("ADC修正(V):"))
        self.sb_offs = QDoubleSpinBox(); self.sb_offs.setRange(-1.0, 1.0); self.sb_offs.setSingleStep(0.001)
        self.sb_offs.setDecimals(3); self.sb_offs.setValue(0.0); self.sb_offs.setFixedWidth(100)
        self.sb_offs.valueChanged.connect(lambda v: (setattr(self, '_offs', v), self._save()))
        cr.addWidget(self.sb_offs)
        # 偏移视觉调整 (per-channel) — 放在 ADC修正 这一行后面
        cr.addStretch(); self._off_row_L = cr   # save ref for dynamic rebuild
        self._off_dynamic_widgets = []
        o.addLayout(cr)

        # action row
        ar = QHBoxLayout()
        self.lbl_st = QLabel("就绪"); self.lbl_st.setStyleSheet("color: gray;"); ar.addWidget(self.lbl_st)
        ar.addSpacing(12)
        self.btn_clear = QPushButton("清除波形")
        self.btn_clear.clicked.connect(self._reset); ar.addWidget(self.btn_clear)
        self.btn_clr_cur = QPushButton("清除游标")
        self.btn_clr_cur.clicked.connect(self._hide_cursors); ar.addWidget(self.btn_clr_cur)
        ar.addStretch()
        self.btn_merge = QPushButton("合并显示"); self.btn_merge.setCheckable(True)
        self.btn_merge.toggled.connect(self._toggle_merge); ar.addWidget(self.btn_merge)
        o.addLayout(ar)

        # stacked: [0]=single-independent, [1]=diff-independent, [2]=merged
        self._stack = QStackedWidget()
        sp0, p0, c0, l0 = _build_independent(SE_CH); self._stack.addWidget(sp0)
        self._se_p, self._se_c, self._se_l = p0, c0, l0
        sp1, p1, c1, l1 = _build_independent(DF_CH); self._stack.addWidget(sp1)
        self._df_p, self._df_c, self._df_l = p1, c1, l1
        self._ip, self._ic, self._il = p0, c0, l0
        # merged plot (built first so we can add it to stack)
        self._mp = pg.PlotWidget(viewBox=_VB())
        self._mp.setLabel("left", "电压", units="V"); self._mp.setLabel("bottom", "样本序号")
        self._mp.setYRange(0, 4, padding=0); self._mp.setXRange(0, 1000, padding=0)
        self._mp.enableAutoRange(axis='x', enable=False); self._mp.showGrid(x=True, y=True)
        for i, d in enumerate(SE_CH):
            self._m_cur.append(self._mp.plot(pen=pg.mkPen(color=d["color"], width=1.5), name=d["name"]))
            tx = pg.TextItem(text=d["name"], color=d["color"], anchor=(0, 0.5), fill=pg.mkColor(0, 0, 0, 160))
            tx.setFont(QFont("", 11)); tx.setZValue(10); self._mp.addItem(tx); self._m_txt.append(tx)
        self._stack.addWidget(self._mp)   # merged page = index 2
        self._mp.getViewBox().scene().sigMouseClicked.connect(self._clk_merge)
        self._mp.getViewBox().sigRangeChanged.connect(self._merge_range_chg)
        o.addWidget(self._stack, stretch=1)

    # helpers
    def _mask(self):
        m = 0
        for i, c in enumerate(self._chks):
            if c.isChecked(): m |= 1 << self._ch_defs[i]["phys"]
        return m

    def _sync_names(self):
        for i, e in enumerate(self._edits):
            if i < len(self._ip):
                self._ip[i].setLabel("left", e.text(), units="V"); self._il[i].setText(e.text())
            if i < len(self._m_txt): self._m_txt[i].setText(e.text())

    def _name_chg(self, i, t):
        if i < len(self._ip): self._ip[i].setLabel("left", t, units="V"); self._il[i].setText(t)
        if i < len(self._m_txt): self._m_txt[i].setText(t)

    def _schedule(self): self._update_plot()

    def _prepare_curve(self, curve):
        """Use GL line plot + manual decimation so renders stay thin & fast."""
        curve.setClipToView(True)

    # ─── PLOT UPDATE ───
    def _update_plot(self):
        with self._lock:
            raw_bufs = self._buf
        cmap = self._bmap
        p2d = {d["phys"]: ii for ii, d in enumerate(self._ch_defs)}

        N_SHORT = 2000   # below this: render every point; above: min/max envelope

        def _envelope(volts, n):
            """Build direction-aware min/max envelope for a voltage array."""
            if n <= N_SHORT:
                xx = np.arange(n, dtype=np.float64)
                return xx, volts
            bs = max(n // 4000, 2)
            trim = n - (n % bs)
            v = volts[:trim].reshape(-1, bs)
            v_max = v.max(axis=1); v_min = v.min(axis=1)
            falling = v[:, 0] > v[:, -1]
            x = np.arange(0, trim, bs, dtype=np.float64)
            x2 = np.empty(len(v_max) * 2, dtype=np.float64)
            y2 = np.empty(len(v_max) * 2, dtype=np.float32)
            x2[0::2] = x;      x2[1::2] = x + bs
            y2[0::2] = np.where(falling, v_max, v_min)
            y2[1::2] = np.where(falling, v_min, v_max)
            return x2, y2

        for ci in range(len(self._ip)):
            self._ic[ci].setData([], [])
            if ci < len(self._m_cur): self._m_cur[ci].setData([], [])
            if ci >= len(cmap): continue
            ri = p2d.get(cmap[ci])
            if ri is None or ri >= len(raw_bufs): continue
            b = raw_bufs[ri]
            if not b: continue
            off = self._spins[ri].value() if ri < len(self._spins) else 0.0
            volts = _raw2v(b, self._ch_defs[ri]["diff"]) + np.float32(off)
            xx, yy = _envelope(volts, len(volts))
            self._ic[ci].setData(xx, yy)
            if ci < len(self._m_cur): self._m_cur[ci].setData(xx, yy)

        self._refresh_labels()
        self._refresh_markers()

    def _refresh_markers(self):
        """Keep placed marker readouts in sync with the (possibly offset-shifted) curve."""
        for ix in range(len(self._cur_px)):
            if any(px is not None for px in self._cur_px[ix]):
                self._render_ind_text(ix); self._position_ind_label(ix)
        if any(px is not None for px in self._m_px):
            self._render_merge_text()

    def _refresh_labels(self):
        for i, pw in enumerate(self._ip):
            vb = pw.getViewBox(); xr = vb.viewRange()[0]
            xd, yd = self._ic[i].getData()
            if xd is None or not len(xd): self._il[i].setPos(-9999, -9999); continue
            idx = next((j for j in range(0, len(xd), max(1, len(xd) // 200)) if xd[j] >= xr[0]), None)
            if idx is None or xd[idx] > xr[1]: self._il[i].setPos(-9999, -9999)
            else: self._il[i].setPos(xd[idx] + 2, yd[idx])
        vb = self._mp.getViewBox(); xr = vb.viewRange()[0]
        for i, cv in enumerate(self._m_cur):
            xd, yd = cv.getData()
            if xd is None or not len(xd): self._m_txt[i].setPos(-9999, -9999); continue
            idx = next((j for j in range(0, len(xd), max(1, len(xd) // 200)) if xd[j] >= xr[0]), None)
            if idx is None or xd[idx] > xr[1]: self._m_txt[i].setPos(-9999, -9999)
            else: self._m_txt[i].setPos(xd[idx] + 2, yd[idx])

    # ─── CURSOR independent ───
    def _volts_at(self, bufs, ri, px):
        """Displayed voltage at sample px for ch_def ri (matches the drawn curve)."""
        b = bufs[ri]
        off = self._spins[ri].value() if ri < len(self._spins) else 0.0
        return _raw2v(b[px], self._ch_defs[ri]["diff"]) + np.float32(off)

    def _clear_cursors_ind(self, ix):
        for m in range(2):
            self._cur_vl[ix][m].setVisible(False)
            self._cur_tx[ix][m].setVisible(False)
            self._cur_px[ix][m] = None
            self._cur_pos[ix][m] = None
        self._cur_next[ix] = 0

    def _clear_cursors_merge(self):
        for ci in range(len(self._m_cur_vl)):
            for m in range(2):
                self._m_cur_vl[ci][m].setVisible(False)
                self._m_cur_tx[ci][m].setVisible(False)
        self._m_px = [None, None]; self._m_next = 0; self._m_clk = [None, None]

    def _clk_ind(self, ev, ix):
        pw = self._ip[ix]; vb = pw.getViewBox()
        if not vb.sceneBoundingRect().contains(ev.scenePos()): return
        if ev.button() != Qt.MouseButton.LeftButton:          # right-click clears this plot's markers
            self._clear_cursors_ind(ix); return
        spos = ev.scenePos()
        px = int(round(vb.mapSceneToView(spos).x()))
        with self._lock:
            bufs = [list(b) for b in self._buf]
        cmap = self._bmap
        p2d = {d["phys"]: ii for ii, d in enumerate(self._ch_defs)}
        if ix >= len(cmap): return
        ri = p2d.get(cmap[ix])
        if ri is None or ri >= len(bufs) or not bufs[ri]: return
        if px < 0 or px >= len(bufs[ri]): return
        m = self._cur_next[ix]; self._cur_next[ix] = 1 - m      # toggle ① <-> ②
        self._cur_px[ix][m] = px
        self._cur_pos[ix][m] = spos
        self._cur_vl[ix][m].setVisible(True); self._cur_vl[ix][m].setPos(px)
        self._render_ind_text(ix)
        self._position_ind_label(ix)

    def _render_ind_text(self, ix):
        """Refresh both marker text boxes for plot ix (adds ΔV when both set)."""
        cmap = self._bmap
        p2d = {d["phys"]: ii for ii, d in enumerate(self._ch_defs)}
        if ix >= len(cmap): return
        ri = p2d.get(cmap[ix])
        if ri is None: return
        nm = self._edits[ri].text() if ri < len(self._edits) else self._ch_defs[ri]["pin"]
        with self._lock:
            bufs = [list(b) for b in self._buf]
        if not bufs[ri]: return
        volts = [None, None]
        for m in range(2):
            px = self._cur_px[ix][m]
            if px is None or px >= len(bufs[ri]):
                self._cur_tx[ix][m].setVisible(False); continue
            volts[m] = float(self._volts_at(bufs, ri, px))
            self._cur_tx[ix][m].setVisible(True)
        both = volts[0] is not None and volts[1] is not None
        for m in range(2):
            px = self._cur_px[ix][m]
            if px is None or volts[m] is None: continue
            t = f"{MK_LABELS[m]} {nm} #{px}\nV={volts[m]:+.3f}V"
            if both and m == 1:
                t += f"\nΔV={volts[1]-volts[0]:+.3f}V Δn={self._cur_px[ix][1]-self._cur_px[ix][0]:+d}"
            self._cur_tx[ix][m].setText(t)

    def _position_ind_label(self, ix):
        vb = self._ip[ix].getViewBox()
        for m in range(2):
            if self._cur_pos[ix][m] is not None and self._cur_tx[ix][m].isVisible():
                loc = vb.mapSceneToView(self._cur_pos[ix][m])
                self._cur_tx[ix][m].setPos(loc.x() + 5, loc.y())

    def _ind_range_chg(self):
        for i in range(len(self._ip)):
            self._position_ind_label(i)
            self._follow_name(i)

    def _follow_name(self, i):
        pw = self._ip[i]; vb = pw.getViewBox(); xr = vb.viewRange()[0]
        xd, yd = self._ic[i].getData()
        if xd is None or not len(xd): self._il[i].setPos(5, 2.0); return
        idx = next((j for j in range(0, len(xd), max(1, len(xd) // 200)) if xd[j] >= xr[0]), None)
        if idx is None or xd[idx] > xr[1]: self._il[i].setPos(-9999, -9999)
        else: self._il[i].setPos(xd[idx] + 2, yd[idx])

    def _clk_merge(self, ev):
        vb = self._mp.getViewBox()
        if not vb.sceneBoundingRect().contains(ev.scenePos()): return
        if ev.button() != Qt.MouseButton.LeftButton:       # right-click clears both markers
            self._clear_cursors_merge(); return
        spos = ev.scenePos()
        px = int(round(vb.mapSceneToView(spos).x()))
        m = self._m_next; self._m_next = 1 - m               # toggle ① <-> ②
        self._m_px[m] = px; self._m_clk[m] = spos
        self._render_merge_text()

    def _render_merge_text(self):
        cmap = self._bmap
        p2d = {d["phys"]: ii for ii, d in enumerate(self._ch_defs)}
        with self._lock:
            bufs = [list(b) for b in self._buf]
        both = self._m_px[0] is not None and self._m_px[1] is not None
        for m in range(2):
            px = self._m_px[m]
            for ci in range(len(self._m_cur)):
                if ci >= len(cmap) or px is None:
                    self._m_cur_vl[ci][m].setVisible(False)
                    self._m_cur_tx[ci][m].setVisible(False); continue
                i = p2d.get(cmap[ci])
                if i is None or i >= len(bufs) or not bufs[i] or px < 0 or px >= len(bufs[i]):
                    self._m_cur_vl[ci][m].setVisible(False)
                    self._m_cur_tx[ci][m].setVisible(False); continue
                nm = self._edits[i].text() if i < len(self._edits) else self._ch_defs[i]["pin"]
                v = float(self._volts_at(bufs, i, px))
                self._m_cur_vl[ci][m].setVisible(True); self._m_cur_vl[ci][m].setPos(px)
                t = f"{MK_LABELS[m]} {nm} #{px}\nV={v:+.3f}V"
                if both and m == 1:
                    px0 = self._m_px[0]
                    if px0 < len(bufs[i]):
                        v0 = float(self._volts_at(bufs, i, px0))
                        t += f"\nΔV={v-v0:+.3f}V Δn={px-px0:+d}"
                self._m_cur_tx[ci][m].setText(t)
                self._m_cur_tx[ci][m].setVisible(True)
        self._position_merge_text()

    def _position_merge_text(self):
        vb = self._mp.getViewBox()
        for m in range(2):
            if self._m_clk[m] is None: continue
            pt = vb.mapSceneToView(self._m_clk[m])
            p_d = vb.mapSceneToView(self._m_clk[m] + QPointF(0, 18))
            dy = abs(p_d.y() - pt.y()); idx = 0
            for ci in range(len(self._m_cur)):
                if ci < len(self._m_cur_tx) and self._m_cur_tx[ci][m].isVisible():
                    self._m_cur_tx[ci][m].setPos(pt.x() + 5, pt.y() - idx * dy)
                    idx += 1

    def _merge_range_chg(self):
        self._position_merge_text()

    def _hide_cursors(self):
        for ix in range(len(self._cur_vl)):
            for m in range(2):
                self._cur_vl[ix][m].setVisible(False)
                self._cur_tx[ix][m].setVisible(False)
            if ix < len(self._cur_px): self._cur_px[ix] = [None, None]
            if ix < len(self._cur_pos): self._cur_pos[ix] = [None, None]
            if ix < len(self._cur_next): self._cur_next[ix] = 0
        for ci in range(len(self._m_cur_vl)):
            for m in range(2):
                self._m_cur_vl[ci][m].setVisible(False)
                self._m_cur_tx[ci][m].setVisible(False)
        self._m_px = [None, None]; self._m_next = 0; self._m_clk = [None, None]

    def _toggle_merge(self, ck):
        if ck:
            self._stack.setCurrentIndex(2); self.btn_merge.setText("独立显示"); self._hide_cursors()
        else:
            self._stack.setCurrentIndex(0 if self._mode == MODE_SINGLE else 1)
            self.btn_merge.setText("合并显示"); self._hide_cursors()

    # ─── BURST CAPTURE ───
    def _go(self):
        if not self.link.is_open():
            self.log_signal.emit("[WARN] 未连接"); return
        mask = self._mask()
        if not mask: self.log_signal.emit("[WARN] 至少选择一个通道"); return
        rate = self.sb_rate.value(); ns = self.sb_pts.value()
        ch_map = sorted(set(self._ch_defs[i]["phys"] for i, c in enumerate(self._chks) if c.isChecked()))
        nb = len(ch_map)
        with self._lock:
            for b in self._buf: b.clear()
        self._hide_cursors()                                    # new waveform -> clear old markers
        self._bs += 1; self._bs_ack = self._bs
        self._bpend = True; self._bexp = ns * nb; self._brec = 0
        self._bnb = nb; self._bmask = mask; self._bmap = ch_map; self._bbuf = [[] for _ in range(nb)]
        self.btn_go.setEnabled(False)
        pins = ", ".join(self._ch_defs[i]["pin"] for i, c in enumerate(self._chks) if c.isChecked())
        self.lbl_st.setText(f"采集中... {ns}样本 x {nb}通道({pins}) @ {rate}Hz")
        self.lbl_st.setStyleSheet("color: orange; font-weight: bold;")
        self.link.send(build_adc_config(mask, rate, mode=self._mode).to_bytes())
        QTimer.singleShot(50, lambda: self._send_burst(mask, ns)); self._btimer.start(5000)

    def _send_burst(self, mask, ns):
        self.link.send(build_adc_burst(mask, ns).to_bytes())
        self.log_signal.emit(f"[TX] BURST ch=0x{mask:02X} n={ns} rate={self.sb_rate.value()}Hz mode={self._mode}")

    def _reset(self):
        self._bpend = False; self._bs_ack = 0; self._btimer.stop(); self.btn_go.setEnabled(True)
        with self._lock:
            for b in self._buf: b.clear()
        for c in self._ic: c.setData([], [])
        for c in self._m_cur: c.setData([], [])
        self._hide_cursors(); self.lbl_st.setText("已清除"); self.lbl_st.setStyleSheet("color: gray;")
        self.log_signal.emit("[INFO] ADC波形已清除")

    def on_frame(self, frame):
        if frame.cmd == CMD_ADC_DATA:
            try:
                _, ch_mask, mode, raw = parse_adc_data(frame.payload)
            except ValueError: return
            samples = np.frombuffer(raw, dtype=np.uint16)
            if not len(samples): return
            ne = bin(ch_mask).count("1")
            if not ne: return
            spc = len(samples) // ne
            if not spc: return
            samples = samples[:spc * ne]; rs = samples.reshape(spc, ne)
            if self._bpend and ch_mask == self._bmask and self._bs == self._bs_ack:
                self._btimer.start(5000)
                buf = self._bbuf
                for c in range(ne):
                    buf[c].append(rs[:, c].copy())
                self._brec += spc; exp = self._bexp // ne if ne else 0
                if self._brec >= exp:
                    cmap = self._bmap
                    p2d = {d["phys"]: ii for ii, d in enumerate(self._ch_defs)}
                    with self._lock:
                        for col in range(ne):
                            phys = cmap[col] if col < len(cmap) else col
                            idx = p2d.get(phys)
                            if idx is not None:
                                self._buf[idx] = list(np.concatenate(buf[col])) if buf[col] else []
                    self._bpend = False; self.log_signal.emit(f"[INFO] Burst完成: {exp} spc/ch"); self._burst_done.emit()
        elif frame.cmd == CMD_ACK and len(frame.payload) >= 2:
            ok = frame.payload[1]
            if ok != 0:
                self.log_signal.emit(f"[WARN] 命令 0x{frame.payload[0]:02X} 执行失败")

    def _after_capture(self):
        """Post-capture: update plot & export."""
        try: self._update_plot(); self._export_excel()
        except Exception as e: self.log_signal.emit(f"[WARN] 采集异常：{e}")
        self.btn_go.setEnabled(True)
        spc = self._bexp // max(self._bnb, 1)
        self.lbl_st.setText(f"完成: {spc} 样本/通道"); self.lbl_st.setStyleSheet("color: green; font-weight: bold;")
        self._btimer.stop()

    @pyqtSlot()
    def _bust_slot(self):
        self._after_capture()

    def _bust_timeout(self):
        if not self._bpend: return
        if not self._brec:
            self._bpend = False; self.btn_go.setEnabled(True)
            self.lbl_st.setText("采集超时（无数据）"); self.lbl_st.setStyleSheet("color: red; font-weight: bold;")
            self.log_signal.emit("[WARN] 采集超时：未收到任何 ADC 数据"); return
        # write partial data to buffer
        cmap = self._bmap; nb = self._bnb; buf = self._bbuf
        with self._lock:
            for col in range(nb):
                phys = cmap[col] if col < len(cmap) else col
                idx = next((ii for ii, d in enumerate(self._ch_defs) if d["phys"] == phys), None)
                if idx is not None and idx < len(self._buf):
                    self._buf[idx] = list(np.concatenate(buf[col])) if buf[col] else []
        self._bpend = False
        self._after_capture()

    # ─── EXCEL EXPORT ───
    def _export_excel(self):
        ld = r"D:\test\STM32H723ZGT6\host\adc_logs"; os.makedirs(ld, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self._save_excel(os.path.join(ld, "adc_latest.xlsx")); self.log_signal.emit(f"[INFO] Excel已保存")

    def _export_as(self):
        ld = r"D:/test/STM32H723ZGT6/host/adc_logs"; os.makedirs(ld, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        p, _ = QFileDialog.getSaveFileName(self, "另存为", os.path.join(ld, f"adc_{ts}.xlsx"), "Excel (*.xlsx)")
        if p: self._save_excel(p)

    def _save_excel(self, path):
        cmap = self._bmap
        if not cmap: self.log_signal.emit("[WARN] 没有可导出的波形数据"); return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "ADC"
        hdr = ["样本序号"]
        for ci in cmap:
            dff = self._ch_defs[ci]["diff"]; nm = self._edits[ci].text() if ci < len(self._edits) else self._ch_defs[ci]["pin"]
            hdr.append(f"{nm} {'差分' if dff else ''}电压(V)")
        for ci, h in enumerate(hdr, 1): ws.cell(row=1, column=ci, value=h)
        with self._lock: bufs = [list(b) for b in self._buf]
        spc = min(len(bufs[ci]) if ci < len(bufs) else 0 for ci in cmap)
        for r in range(spc):
            ws.cell(row=r + 2, column=1, value=r)
            for ji, ci in enumerate(cmap):
                raw = bufs[ci][r] if ci < len(bufs) and r < len(bufs[ci]) else 0
                ws.cell(row=r + 2, column=ji + 2, value=round(_raw2v(raw, self._ch_defs[ci]["diff"]) - self._offs, 4))
        wb.save(path); self.log_signal.emit(f"[INFO] Excel已保存: {path}")