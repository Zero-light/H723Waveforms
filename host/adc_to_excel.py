#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Direct ADC burst → Excel exporter.  No GUI needed.

Usage:
    python adc_to_excel.py COM7                      # PA6 only, 20kHz, 1000 samples
    python adc_to_excel.py COM7 --ch1                # PA6 + PA7
    python adc_to_excel.py COM7 --ch1 --ch2          # PA6 + PA7 + PC1
    python adc_to_excel.py COM7 --ch1 --ch2 --ch3    # + CLK(PB0)
    python adc_to_excel.py COM7 --ch1 --ch2 --ch3 --ch4  # + XYNC(PC4), all 5
    python adc_to_excel.py COM7 --ch1 --rate 100000 --samples 8192
    python adc_to_excel.py COM7 --snap               # snapshot: PA6+CLK+XYNC, raw output
"""

import argparse
import os
import struct
import time
import sys
import serial
import openpyxl
from datetime import datetime

from comm.protocol import (
    build_adc_config, build_adc_burst,
    unpack_frame, CMD_ADC_DATA,
)

ADC_VREF   = 3.3
ADC_MAX    = 4095.0
CH_PINS    = ["PA6", "PA7", "PC1", "CLK", "XYNC"]


def main():
    parser = argparse.ArgumentParser(description="ADC burst → Excel")
    parser.add_argument("port", help="COM port, e.g. COM7")
    parser.add_argument("--ch1", action="store_true", help="Enable PA7")
    parser.add_argument("--ch2", action="store_true", help="Enable PC1")
    parser.add_argument("--ch3", action="store_true", help="Enable CLK (PB0)")
    parser.add_argument("--ch4", action="store_true", help="Enable XYNC (PC4)")
    parser.add_argument("--snap", action="store_true",
                        help="Snapshot mode: PA6+CLK+XYNC, raw ADC output")
    parser.add_argument("--rate", type=int, default=20000,
                        help="Sample rate Hz (default 20000)")
    parser.add_argument("--samples", type=int, default=1000,
                        help="Samples per channel (default 1000, max 8192)")
    parser.add_argument("--output", default="",
                        help="Output .xlsx path (default: auto-timestamp)")
    parser.add_argument("--offset", type=float, default=0.0,
                        help="ADC voltage offset correction (default 0.0)")
    parser.add_argument("--raw", action="store_true",
                        help="Export raw ADC values instead of voltage")
    args = parser.parse_args()

    if args.snap:
        ch_mask = 0x01  # PA6
        ch_mask |= 0x08  # CLK (PB0)
        ch_mask |= 0x10  # XYNC (PC4)
        raw_output = True
    else:
        ch_mask = 0x01  # PA6 always on
        if args.ch1:
            ch_mask |= 0x02   # PA7
        if args.ch2:
            ch_mask |= 0x04   # PC1
        if args.ch3:
            ch_mask |= 0x08   # CLK (PB0)
        if args.ch4:
            ch_mask |= 0x10   # XYNC (PC4)
        raw_output = args.raw

    num_ch = bin(ch_mask).count("1")
    active_pins = [
        CH_PINS[i] for i in range(len(CH_PINS)) if ch_mask & (1 << i)]

    print(f"[INFO] Opening {args.port} ...")
    ser = serial.Serial(args.port, baudrate=115200, timeout=2.0)
    time.sleep(0.5)
    ser.read(ser.in_waiting)  # flush stale

    # ── Send config ────────────────────────────────────────────────
    cfg = build_adc_config(ch_mask, args.rate, mode=0)
    ser.write(cfg.to_bytes())
    time.sleep(0.1)

    # ── Send burst ─────────────────────────────────────────────────
    frame = build_adc_burst(ch_mask, args.samples)
    ser.write(frame.to_bytes())
    print(f"[TX] BURST ch=0x{ch_mask:02X} samples={args.samples} "
          f"rate={args.rate}Hz")

    # ── Receive all CMD_ADC_DATA frames ────────────────────────────
    buf_raw = b""
    burst_total = args.samples * num_ch
    received    = 0
    ch_buffers   = [[] for _ in range(num_ch)]

    print("[RX] Waiting for data frames...")
    while received < burst_total:
        raw = ser.read(4096)
        if not raw:
            if received > 0:
                break  # timeout but we have data
            print("[WARN] Serial timeout — no data received")
            ser.close()
            return 1
        buf_raw += raw

        while True:
            frame, buf_raw = unpack_frame(buf_raw)
            if frame is None:
                break
            if frame.cmd == CMD_ADC_DATA:
                payload = frame.payload
                if len(payload) < 4:
                    continue
                seq  = payload[0] | (payload[1] << 8)
                mask = payload[2]
                raw_bytes = payload[4:]
                samples = struct.unpack(
                    "<" + "H" * (len(raw_bytes) // 2), raw_bytes)
                n = len(samples)
                received += n
                for i in range(0, n, num_ch):
                    for c in range(num_ch):
                        if i + c < len(samples):
                            ch_buffers[c].append(samples[i + c])
                pct = min(100, received * 100 // burst_total)
                print(f"\r[RX] {received}/{burst_total} ({pct}%)  "
                      f"seq={seq}", end="", flush=True)

    print()
    ser.close()

    if received == 0:
        print("[ERROR] No data received.")
        return 1

    # Trim to requested size
    for c in range(num_ch):
        ch_buffers[c] = ch_buffers[c][:args.samples]

    # ── Write Excel ────────────────────────────────────────────────
    out_path = args.output
    if not out_path:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        log_dir = os.path.join(script_dir, "adc_logs")
        os.makedirs(log_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = "snap" if args.snap else "adc_burst"
        out_path = os.path.join(log_dir, f"{prefix}_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ADC数据"

    if raw_output:
        headers = ["样本序号"] + [f"{p}(raw)" for p in active_pins]
    else:
        headers = ["样本序号"] + [f"{p}电压(V)" for p in active_pins]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    for r in range(args.samples):
        ws.cell(row=r + 2, column=1, value=r)
        for c in range(num_ch):
            val = ch_buffers[c][r]
            if raw_output:
                ws.cell(row=r + 2, column=c + 2, value=int(val))
            else:
                ws.cell(row=r + 2, column=c + 2,
                        value=round(val * ADC_VREF / ADC_MAX
                                    - args.offset, 4))

    wb.save(out_path)
    print(f"[OK]  Saved: {out_path}  "
          f"({args.samples} samples × {num_ch} channels)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
